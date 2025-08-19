"""PyQt5 CCMI app: mic → Whisper ASR → GPT (translate/brief) → TTS.

Sessions: One-way / Two-party / Two-party+Audience. Briefs + term list + recent context guide translation. Device picker & meters, swap languages, voice test, review table, CSV/XLSX term import, XLSX export.

Online (OpenAI). API key stays in memory only; no disk persistence (temp audio cleaned).
Deps: PyQt5, numpy, sounddevice, soundfile, openai, colorlog; optional: openpyxl.
"""

# stdlib
import sys
import os
import csv
import json
import re
import tempfile
import warnings
import logging
from pathlib import Path

def resource_path(name: str) -> str:
    """
    Resolve an asset path for both:
      • dev: files next to ccmi.py (or in ./resources/)
      • PyInstaller: files extracted to sys._MEIPASS
    """
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    candidates = [
        base / name,
        base / "resources" / name,
        Path.cwd() / name,
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    logging.getLogger().warning("Asset not found: %s (searched: %s)", name, candidates)
    return str(base / name)


# third-party
import numpy as np
import sounddevice as sd
import soundfile as sf
import colorlog
from openai import OpenAI

# PyQt5
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QTextEdit, QPushButton, QComboBox,
    QProgressBar, QDialog, QDialogButtonBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QGroupBox, QFrame, QGraphicsDropShadowEffect,
    QFileDialog, QTabWidget, QShortcut, QAbstractItemView,
    QFormLayout, QRadioButton, QCheckBox, QGridLayout, QStackedLayout, QButtonGroup,
    QSizePolicy, QListWidget, QListWidgetItem
)
from PyQt5.QtCore import (
    QThread, pyqtSignal, QTimer, Qt, QEvent, QRectF, QRect,
    pyqtProperty, QEasingCurve, QPoint, QSize, QUrl
)
from PyQt5.QtGui import (
    QColor, QKeySequence, QPainter, QPainterPath, QPixmap, QLinearGradient,
    QPen, QDesktopServices, QRegion, QFontDatabase, QFontMetrics, QFont, QPalette, QIcon
)

# Logging: colored INFO+; suppress PyQt/OpenAI/http noise and deprecation spam.
warnings.filterwarnings("ignore", category=DeprecationWarning)

handler = colorlog.StreamHandler(stream=sys.stdout)
handler.setFormatter(colorlog.ColoredFormatter(
    "%(log_color)s%(message)s",
    log_colors={
        "DEBUG":    "cyan",
        "INFO":     "green",
        "WARNING":  "yellow",
        "ERROR":    "red",
        "CRITICAL": "red,bg_white",
    },
))

root = logging.getLogger()
root.handlers.clear()
root.addHandler(handler)
root.setLevel(logging.INFO)

for name in ("openai", "openai._base_client", "httpx", "httpcore"):
    lg = logging.getLogger(name)
    lg.setLevel(logging.WARNING)
    lg.propagate = False


SAMPLE_RATE = 16000  # Whisper expects 16 kHz; timers/block sizes derive from this.


# UI label → TTS voice id.
VOICE_CHOICES = [
    ("neutral, balanced",  "alloy"),
    ("warm, natural",      "ash"),
    ("deep, rich",         "ballad"),
    ("bright, expressive", "coral"),
    ("clear, energetic",   "echo"),
    ("calm, steady",       "sage"),
    ("light, youthful",    "shimmer"),
    ("bold, dramatic",     "verse"),
]


def _resample_mono(x: np.ndarray, sr_from: int, sr_to: int) -> np.ndarray:
    """Linear-interp resample to target Hz, returning int16 with clipping.

    Args:
      x: Mono PCM (int16/float) array.
      sr_from: Input sample rate (Hz).
      sr_to: Output sample rate (Hz).
    Returns:
      int16 PCM in [-32768, 32767].
    """
    if sr_from == sr_to:
        return x.astype(np.int16, copy=False)
    xf = x.astype(np.float32)
    dur = len(xf) / sr_from
    t_old = np.linspace(0.0, dur, num=len(xf), endpoint=False, dtype=np.float32)
    t_new = np.linspace(0.0, dur, num=int(round(dur * sr_to)), endpoint=False, dtype=np.float32)
    y = np.interp(t_new, t_old, xf)
    return np.clip(y, -32768, 32767).astype(np.int16)


# OpenAI client helpers: session-only key; no persistence to disk.
_SESSION_API_KEY = ""

def get_api_key() -> str:
    """Return session API key held in memory."""
    return _SESSION_API_KEY

def set_api_key(k: str):
    """Set session API key; trimmed; overrides previous value."""
    global _SESSION_API_KEY
    _SESSION_API_KEY = (k or "").strip()

def new_client():
    """Create OpenAI client or raise if key missing (fast-fail)."""
    key = get_api_key()
    if not key:
        raise RuntimeError("No OpenAI API key set.")
    return OpenAI(api_key=key)


class VolumeMonitorThread(QThread):
    """Mic RMS → 0–100% meter about every 100 ms.

    Emits:
      volume_signal: Int level (0..100).
    """
    volume_signal = pyqtSignal(int)
    
    def __init__(self, input_device):
        super().__init__()
        self.input_device = input_device
        self.running = True
    
    def run(self):
        """Poll input (int16), compute RMS, scale by full-scale int16."""
        def callback(indata, frames, time, status):
            if status:
                pass
            f = indata.astype(np.float32)
            rms = float(np.sqrt(np.mean(np.square(f))))
            volume = int(max(0.0, min(1.0, rms / 32768.0)) * 100)
            self.volume_signal.emit(volume)
        
        with sd.InputStream(samplerate=SAMPLE_RATE,
                            channels=1,
                            dtype='int16',
                            device=self.input_device,
                            callback=callback):
            while self.running:
                sd.sleep(100)
    
    def stop(self):
        """Stop thread and join to ensure clean shutdown."""
        self.running = False
        self.wait()


class RecordingThread(QThread):
    """Capture raw session audio in 100 ms blocks; no VAD/chunking.

    Emits:
      progress_signal: Status text.
      recorded_signal: Concatenated int16 bytes on stop.
    """
    progress_signal = pyqtSignal(str)
    recorded_signal = pyqtSignal(bytes)

    def __init__(self, input_device):
        super().__init__()
        self.input_device = input_device
        self.running = False

    def run(self):
        """Use RawInputStream; report overflow; buffer until stopped."""
        self.running = True
        self.progress_signal.emit("Recording… Speak now.")

        frame_size = int(SAMPLE_RATE * 0.10)  # 100ms blocks
        recorded_frames = []
        try:
            with sd.RawInputStream(
                samplerate=SAMPLE_RATE,
                channels=1,
                dtype='int16',
                blocksize=frame_size,
                device=self.input_device
            ) as stream:
                while self.running:
                    frame, overflowed = stream.read(frame_size)
                    if overflowed:
                        self.progress_signal.emit("Audio overflow occurred!")
                    recorded_frames.append(frame)
        except Exception as e:
            self.progress_signal.emit(f"Recording Error: {e}")
            return

        audio_bytes = b"".join(recorded_frames)
        self.progress_signal.emit("Recording finished.")
        self.recorded_signal.emit(audio_bytes)

    def stop(self):
        """Stop thread and join."""
        self.running = False
        self.wait()


class TranscriptionThread(QThread):
    """Write temp FLAC and call Whisper; clean up file on exit.

    Emits:
      progress_signal: Status text.
      transcribed_signal: Transcript string.
    """
    progress_signal = pyqtSignal(str)
    transcribed_signal = pyqtSignal(str)

    def __init__(self, audio_bytes: bytes, tag: str = "Session", model: str = "whisper-1"):
        super().__init__()
        self.audio_bytes = audio_bytes or b""
        self.tag = tag
        self.model = model

    def run(self):
        """Fast-fail on empty audio; log duration; safe resp.text extraction."""
        prefix = f"[{self.tag}] "
        try:
            if not self.audio_bytes:
                self.progress_signal.emit(prefix + "Transcription Error: empty audio buffer.")
                return

            num_samples = len(self.audio_bytes) // 2  # int16 = 2 bytes
            duration_s = num_samples / float(SAMPLE_RATE)
            self.progress_signal.emit(prefix + f"Starting transcription… buffer={len(self.audio_bytes)} bytes (~{duration_s:.2f}s @ {SAMPLE_RATE} Hz mono, int16)")

            audio_array = np.frombuffer(self.audio_bytes, dtype=np.int16).reshape(-1, 1)
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".flac", delete=False) as tmp:
                    tmp_path = tmp.name
                sf.write(tmp_path, audio_array, SAMPLE_RATE)
                self.progress_signal.emit(prefix + "Audio encoded to FLAC; calling Whisper…")

                client = new_client()
                with open(tmp_path, "rb") as f:
                    resp = client.audio.transcriptions.create(
                        model=self.model,
                        file=f
                    )
            finally:
                if tmp_path:
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass

            text = resp.text.strip() if hasattr(resp, "text") else str(resp).strip()
            self.progress_signal.emit(prefix + f"Transcription complete ({len(text)} chars).")
            self.transcribed_signal.emit(text)

        except Exception as e:
            self.progress_signal.emit(prefix + f"Transcription Error: {e}")


class TranslationThread(QThread):
    """GPT translation only; supports verbosity, terminology, and context blocks.

    Emits:
      progress_signal: Status text.
      translated_signal: (source_text, translated_text).
    """
    progress_signal = pyqtSignal(str)
    translated_signal = pyqtSignal(str, str)  # (source_text, translated_text)

    def __init__(self, transcript: str, source_lang: str, target_lang: str,
                 instructions: str, gpt_model: str, context_translations: str = "",
                 term_lines: str = "", verbosity: str = "low"):
        super().__init__()
        self.transcript = transcript
        self.source_lang = source_lang
        self.target_lang = target_lang
        self.instructions = instructions
        self.gpt_model = gpt_model
        self.context_translations = context_translations
        self.term_lines = term_lines
        self.verbosity = (verbosity or "low").strip().lower()

    def run(self):
        """Constructs strict 'translation only' prompt; returns text content."""
        self.progress_signal.emit("Translating text with GPT...")
        try:
            vb = self.verbosity
            if vb == "low":
                vb_rules = "Aim for maximum brevity while natural. Avoid redundant words."
            elif vb in ("normal", "medium"):
                vb_rules = "Be concise and natural."
            elif vb == "high":
                vb_rules = "Be natural and complete; do not add commentary."
            else:
                vb_rules = "Be concise and natural."

            prompt_lines = [
                f"Translate the following source text from {self.source_lang} to {self.target_lang}.",
                "Rules:",
                f"- Follow these style/instruction notes: {self.instructions or '(none)'}",
                f"- Style guidance: {vb_rules}",
                "- Return only the translation (no extra commentary).",
            ]
            if self.term_lines.strip():
                prompt_lines.append("Terminology list (one per line; enforce exact choices):")
                prompt_lines.append(self.term_lines)
            if self.context_translations.strip():
                prompt_lines.append("Context for consistency (do not translate this block):")
                prompt_lines.append(self.context_translations)
            prompt_lines.append(f"Source Text:\n{self.transcript}")
            prompt = "\n".join(prompt_lines)

            client = new_client()
            model_name = (self.gpt_model or "").strip()
            if not model_name:
                raise RuntimeError("No GPT model specified (GPT Model textbox is empty).")

            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are a Customized Consecutive Interpreter."},
                    {"role": "user", "content": prompt},
                ],
            )

            translated_text = (response.choices[0].message.content or "").strip()
            self.progress_signal.emit("Translation complete.")
            self.translated_signal.emit(self.transcript, translated_text)
        except Exception as e:
            self.progress_signal.emit(f"Translation Error: {e}")


class TTSThread(QThread):
    """Generate TTS WAV and stream to output in ~20 ms blocks; emit output level.

    Emits:
      progress_signal: Status text.
      level_signal: Int level (0..100) during playback.
    """
    progress_signal = pyqtSignal(str)
    level_signal = pyqtSignal(int)  # 0..100 output level while playing

    def __init__(self, translated_text: str, voice_name: str, output_device: int):
        super().__init__()
        self.translated_text = translated_text
        self.voice_name = voice_name
        self.output_device = output_device

    def run(self):
        """Request streaming TTS, write temp WAV, play, then clean up."""
        self.progress_signal.emit("Generating TTS from translated text...")
        try:
            client = new_client()

            fd, tts_audio_path = tempfile.mkstemp(suffix=".wav"); os.close(fd)
            with client.audio.speech.with_streaming_response.create(
                model="gpt-4o-mini-tts",
                voice=self.voice_name,
                input=self.translated_text,
                response_format="wav"
            ) as r:
                r.stream_to_file(tts_audio_path)

            data, fs = sf.read(tts_audio_path, dtype='float32', always_2d=True)
            try:
                os.remove(tts_audio_path)
            except Exception:
                pass

            block = int(fs * 0.02)  # 20 ms chunks
            idx = 0
            self.progress_signal.emit("TTS audio playing...")

            with sd.OutputStream(
                samplerate=fs,
                channels=data.shape[1],
                dtype='float32',
                device=self.output_device
            ) as stream:
                while idx < len(data):
                    chunk = data[idx: idx + block]
                    if len(chunk) == 0:
                        break
                    stream.write(chunk)
                    rms = float(np.sqrt(np.mean(np.square(chunk))))
                    level = int(max(0.0, min(1.0, rms)) * 100)
                    self.level_signal.emit(level)
                    idx += block

            self.level_signal.emit(0)
            self.progress_signal.emit("TTS playback finished.")
        except Exception as e:
            self.progress_signal.emit(f"TTS Error: {e}")
            self.level_signal.emit(0)

class BriefFillingThread(QThread):
    """Summarize spoken text into a minimal brief JSON per session mode.

    Emits:
      progress_signal: Status text.
      filled_signal: Dict with fields matching the selected mode.
    """
    progress_signal = pyqtSignal(str)
    filled_signal = pyqtSignal(dict)  # emits a dict with keys per session mode

    def __init__(self, transcript: str, mode: str, gpt_model: str, verbosity: str = "low"):
        super().__init__()
        self.transcript = (transcript or "").strip()
        self.mode = mode  # "one" | "two" | "two_aud"
        self.gpt_model = gpt_model
        self.verbosity = (verbosity or "low").strip().lower()

    def _schema_for_mode(self) -> str:
        """Return JSON schema string for the selected mode."""
        if self.mode == "one":
            return json.dumps({
                "audience_info": "",
                "audience_expectation": "",
                "st": {"purpose": "", "speaker": "", "tone": "", "expectation": "", "notes": ""}
            }, indent=2)
        if self.mode == "two":
            return json.dumps({
                "st": {"purpose": "", "speaker": "", "tone": "", "expectation": "", "notes": ""},
                "ts": {"purpose": "", "speaker": "", "tone": "", "expectation": "", "notes": ""}
            }, indent=2)
        return json.dumps({  # two_aud
            "audience_info": "",
            "audience_expectation": "",
            "st": {"purpose": "", "speaker": "", "tone": "", "expectation": "", "notes": ""},
            "ts": {"purpose": "", "speaker": "", "tone": "", "expectation": "", "notes": ""}
        }, indent=2)

    def run(self):
        """JSON-only output; strip code fences; empty transcript → {}."""
        try:
            self.progress_signal.emit("Brief: summarizing with GPT…")
            if not self.transcript:
                self.filled_signal.emit({})
                return

            vb = self.verbosity
            if vb == "low":
                vb_rules = "Be ultra-concise. Keep each value 2–8 words. No filler. No explanations."
            elif vb in ("normal", "medium"):
                vb_rules = "Be concise. Each value up to ~12 words. No explanations."
            elif vb == "high":
                vb_rules = "Be a bit more descriptive if needed (≤18 words per value). No explanations."
            else:
                vb_rules = "Be concise. No explanations."

            sys_prompt = (
                "You convert a spoken description into a minimal session brief.\n"
                "Return ONLY JSON (no markdown). Use the exact field names from the schema.\n"
                "If a value is unknown, use an empty string.\n"
                f"Verbosity: {vb}. {vb_rules}"
            )
            schema = self._schema_for_mode()
            user_prompt = (
                f"Session mode: {self.mode}\n\n"
                "Fill this schema with concise phrases based ONLY on the speech:\n"
                f"{schema}\n\n"
                "Speech:\n"
                f"\"\"\"\n{self.transcript}\n\"\"\""
            )

            client = new_client()
            model_name = (self.gpt_model or "").strip()
            if not model_name:
                raise RuntimeError("No GPT model specified (GPT Model textbox is empty).")

            resp = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": sys_prompt},
                    {"role": "user", "content": user_prompt},
                ],
            )

            content = (resp.choices[0].message.content or "").strip()
            content = re.sub(r"^```(?:json)?\s*|\s*```$", "", content, flags=re.S)  # strip code fences
            data = json.loads(content)
            if not isinstance(data, dict):
                raise ValueError("Model did not return a JSON object.")
            self.filled_signal.emit(data)
        except Exception as e:
            self.progress_signal.emit(f"Brief error: {e}")
            self.filled_signal.emit({})


class ImageCardButton(QPushButton):
    """Image-backed toggle button with rounded mask, hover scrim, and focus ring.

    Notes:
      - Hover progress is animated (≈160 ms, OutCubic).
      - When checked or hovered, a colored ring is drawn.
      - Label under cursor receives a soft highlight via _text_widgets.
    """
    def __init__(self, *args, bg_path: str = None, radius: int = 20, **kwargs):
        super().__init__(*args, **kwargs)
        self._bg = QPixmap(bg_path) if bg_path else None
        self._radius = radius
        self.setCheckable(True)
        self.setFlat(True)

        self.setMouseTracking(True)
        self._hover_pos = QPoint(-1, -1)
        self._hover_widget = None
        self._hover_progress = 0.0

        from PyQt5.QtCore import QPropertyAnimation
        self._hover_anim = QPropertyAnimation(self, b"hover", self)
        self._hover_anim.setDuration(160)  # ms
        self._hover_anim.setEasingCurve(QEasingCurve.OutCubic)

    def _get_hover(self) -> float:
        return self._hover_progress
    def _set_hover(self, v: float):
        self._hover_progress = max(0.0, min(1.0, float(v)))
        self.update()
    hover = pyqtProperty(float, fget=_get_hover, fset=_set_hover)

    def set_background(self, path: str):
        self._bg = QPixmap(path) if path else None
        self.update()

    def enterEvent(self, e):
        self.setCursor(Qt.PointingHandCursor)
        self._hover_anim.stop()
        self._hover_anim.setStartValue(self._hover_progress)
        self._hover_anim.setEndValue(1.0)
        self._hover_anim.start()
        super().enterEvent(e)

    def leaveEvent(self, e):
        self._hover_widget = None
        self._hover_anim.stop()
        self._hover_anim.setStartValue(self._hover_progress)
        self._hover_anim.setEndValue(0.0)
        self._hover_anim.start()
        super().leaveEvent(e)

    def mouseMoveEvent(self, e):
        self._hover_pos = e.pos()
        old = self._hover_widget
        self._hover_widget = None
        for w in getattr(self, "_text_widgets", []):
            if w.isVisible() and w.geometry().contains(self._hover_pos):
                self._hover_widget = w
                break
        if old is not self._hover_widget:
            self.update()
        super().mouseMoveEvent(e)

    def paintEvent(self, e):
        r = self.rect().adjusted(3, 3, -3, -3)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, True)

        path = QPainterPath()
        path.addRoundedRect(QRectF(r), self._radius, self._radius)
        painter.setClipPath(path)

        # Slight zoom on hover (~+2%).
        scale = 1.0 + 0.02 * self._hover_progress
        if self._bg and not self._bg.isNull():
            target_size = QSize(int(self.width()*scale), int(self.height()*scale))
            scaled = self._bg.scaled(target_size, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            x = (scaled.width()  - self.width())  // 2
            y = (scaled.height() - self.height()) // 2
            painter.drawPixmap(-x, -y, scaled)
        else:
            grad = QLinearGradient(0, 0, 0, self.height())
            grad.setColorAt(0, QColor("#253847"))
            grad.setColorAt(1, QColor("#192a35"))
            painter.fillPath(path, grad)

        if self._hover_progress > 0:
            painter.fillPath(path, QColor(255, 255, 255, int(20 * self._hover_progress)))

        # Full-card scrim (alpha 0..255); lift a bit on hover.
        base_top, base_bot = 160, 195
        lift = int(18 * self._hover_progress)

        full_grad = QLinearGradient(0, 0, 0, self.height())
        full_grad.setColorAt(0.00, QColor(0, 0, 0, base_top + lift))
        full_grad.setColorAt(1.00, QColor(0, 0, 0, base_bot + lift))
        painter.fillPath(path, full_grad)

        # Highlight label under cursor using its text bounds.
        if self._hover_widget is not None:
            w = self._hover_widget
            fm = w.fontMetrics()
            max_w = min(int(self.width()*0.80), max(320, int(w.width()*0.90)))
            br = fm.boundingRect(0, 0, max_w, 1000,
                                 Qt.TextWordWrap | Qt.AlignHCenter,
                                 w.text())
            cx, cy = w.geometry().center().x(), w.geometry().center().y()
            glow_rect = QRectF(int(cx - br.width()/2)-10,
                               int(cy - br.height()/2)-6,
                               br.width()+20, br.height()+12)
            glow_path = QPainterPath()
            glow_path.addRoundedRect(glow_rect, 10, 10)
            painter.fillPath(glow_path, QColor(255, 255, 255, int(22 * self._hover_progress)))

        # Accent ring; thicker when checked.
        ring = QColor("#35e0a7") if (self.isChecked() or self._hover_progress > 0) else QColor("#304354")
        width = 2 if self.isChecked() else (1 + int(1 * self._hover_progress))
        painter.setClipping(False)
        painter.setPen(QPen(ring, width))
        painter.drawPath(path)
        painter.setClipPath(path)

        super().paintEvent(e)


class MainWindow(QWidget):
    """CCMI main window: frameless, custom title bar; fits screen.

    Tabs: Setup, Termlist, Review.
    Setup: pick session (One-way / Two-party / +Audience), fill briefs (Audience/ST/TS),
    choose voices & test, set GPT model/API key, swap Source↔Target.
    Termlist: import/edit term pairs (CSV/XLSX). Review: segment table (copy last, export XLSX).

    Live I/O: device picker (in/out), meters + volume monitor, status chip with pulse,
    record/action timers, toasts.

    Pipeline: Record → ASR (Whisper thread) → GPT translate (uses context + term pairs) → TTS
    (queued; voice follows direction). Also: brief record → ASR → GPT brief-filling.

    Includes shortcuts and clean shutdown of all worker threads.
    """
    def __init__(self, input_device=None, output_device=None):
        super().__init__()
        
        # Create toast early so resizeEvent/geometry ops are always safe.
        self.toast = QLabel("", self)
        self.toast.setObjectName("toast")
        self.toast.hide()
        self._toast_timer = QTimer(self)
        self._toast_timer.setSingleShot(True)
        self._toast_timer.timeout.connect(self._hide_toast)
        
        self.setWindowTitle("Customized Consecutive Machine Interpreter")
        
        # Frameless to avoid native titlebar artifacts; custom drag/titlebar.
        self.setWindowFlag(Qt.FramelessWindowHint, True)
        self._title_h = 30
        self._drag_pos = None
        self._build_title_bar()

        if input_device is None or output_device is None:
            input_device, output_device = self._detect_default_devices()

        self.fixed_input_device = input_device
        self.fixed_output_device = output_device

        self.volume_monitor_thread = None
        self.recording_thread = None
        self.transcription_thread = None
        self.translation_thread = None
        self.tts_thread = None
        self.brief_recording_thread = None
        self.brief_transcription_thread = None
        self.brief_filling_thread = None

        self.test_tts_thread = None
        self.pending_source_text = ""
        self.tts_queue = []
        self.tts_busy = False
        self.tts_timer = None
        self.tts_muted = False
        self.reverse_mode = False  # flips on swap_languages()
        self._tts_context = "none"  # 'translate' or 'hear' (test)
        self._chip_restore = None

        self._status_timer = None
        self._record_timer = None
        self._record_ms = 0
        
        self._action_timer = None  # timer near Idle chip
        self._action_ms = 0

        self._accent = "emerald"
        self._theme_variant = "dim"

        self.init_ui()
        self.apply_dark_theme(self._accent, self._theme_variant)
        self._chip_restore = (self.stage_chip.text(), self.stage_chip.property("state"))
        self.fit_to_work_area()
        self.start_volume_monitor()

        # App-wide shortcuts.
        sc_rec = QShortcut(QKeySequence("Shift+Space"), self)
        sc_rec.setContext(Qt.ApplicationShortcut)
        sc_rec.activated.connect(self.toggle_recording)

        for ks in ("Ctrl+Return", "Ctrl+Enter"):
            sc = QShortcut(QKeySequence(ks), self)
            sc.setContext(Qt.ApplicationShortcut)
            sc.activated.connect(self.swap_languages)

        if sys.platform == "darwin":
            sc_cmd = QShortcut(QKeySequence("Meta+Return"), self)
            sc_cmd.setContext(Qt.ApplicationShortcut)
            sc_cmd.activated.connect(self.swap_languages)

    def _accent_hex(self) -> str:
        """Return current accent as hex for stylesheet usage."""
        return {
            "emerald": "#2ee59d",
            "cyan":    "#39c2ff",
            "violet":  "#b896ff",
            "amber":   "#ffc36b",
        }.get(self._accent, "#2ee59d")

    def _make_card(self, icon_emoji: str, title: str, subtitle: str, desc: str, bg_path: str = None):
        """Create a session card (image background, centered text, hover-aware labels)."""
        btn = ImageCardButton(bg_path=bg_path, radius=20)
        btn.setObjectName("cardToggle")

        v = QVBoxLayout(btn)
        v.setContentsMargins(32, 24, 32, 24)
        v.setSpacing(10)

        e   = QLabel(icon_emoji)
        e.setObjectName("cardEmoji")
        e.setAlignment(Qt.AlignHCenter)

        h1  = QLabel(title)
        h1.setObjectName("cardTitle")
        h1.setAlignment(Qt.AlignHCenter)
        h1.setWordWrap(True)

        sub = QLabel(subtitle)
        sub.setObjectName("cardSubtitle")
        sub.setAlignment(Qt.AlignHCenter)
        sub.setWordWrap(True)

        d   = QLabel(desc)
        d.setObjectName("cardDesc")
        d.setAlignment(Qt.AlignHCenter)
        d.setWordWrap(True)

        for w in (h1, sub, d):
            w.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            w.setContentsMargins(12, 0, 12, 0)
            w.setTextFormat(Qt.PlainText)

        v.addStretch(1)
        for w in (e, h1, sub, d):
            v.addWidget(w, 0, Qt.AlignHCenter)
        v.addStretch(1)

        for lbl in (h1, sub, d):
            eff = QGraphicsDropShadowEffect(btn)
            eff.setBlurRadius(18)
            eff.setOffset(0, 1)
            eff.setColor(QColor(0, 0, 0, 180))
            lbl.setGraphicsEffect(eff)

        btn._text_widgets = [e, h1, sub, d]
        return btn

    def init_ui(self):
        """Build UI; later reparent Idle/timer/session to tab corner to save space.

        Notes:
          - GroupBox titles styled as 'chips' without clipping.
          - Timers use fixed-pitch digits for readability.
        """
        self.setContentsMargins(16, 16, 16, 16)
        img_one     = resource_path("card1.png")
        img_two     = resource_path("card2.png")
        img_two_aud = resource_path("card3.png")

        # GroupBox style: chip-like title, avoids clipping.
        accent_hex = self._accent_hex()
        card_css = f"""
            QGroupBox {{
                background: qlineargradient(x1:0,y1:0, x2:0,y2:1,
                    stop:0 #1e2a36, stop:1 #1a2530);
                border: 1px solid #304354;
                border-radius: 14px;
                margin-top: 26px;
                padding: 16px 14px 14px 14px;
            }}
            QGroupBox:hover {{ border-color: #375064; }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 12px; top: -6px;
                padding: 6px 14px;
                color: {accent_hex};
                font-weight: 900; font-size: 20px; letter-spacing: .3px;
                background: #19232d;
                border: 1px solid #304354;
                border-radius: 10px;
            }}
        """

        root = QVBoxLayout(self)
        root.setSpacing(14)
        root.setContentsMargins(16, self._title_h + 8, 16, 16)

        dev_card = QGroupBox("")
        self._dev_card = dev_card
        dev_card.setObjectName("devCard")
        dev_card.setStyleSheet("""
            #devCard {
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #1e2a36, stop:1 #19252f);
                border: 1px solid #304354;
                border-radius: 14px;
                margin-top: 8px;
                padding: 12px;
            }
            #devCard::title { padding: 0; height: 0; margin: 0; }
        """)

        dev_card.setProperty("stage", "idle")
        dev_col = QVBoxLayout()
        dev_col.setSpacing(10)
        dev_col.setContentsMargins(0, 0, 0, 0)

        self.input_row_frame = QFrame()
        self.input_row_frame.setObjectName("inputRowFrame")
        self.input_row_frame.setStyleSheet("""
            #inputRowFrame { background: #202e3a; border: 1px solid #304354; border-radius: 12px; }
        """)
        self.input_row_frame.setProperty("stage", "idle")
        inp_row = QHBoxLayout(self.input_row_frame)
        inp_row.setContentsMargins(12, 10, 12, 10)
        inp_row.setSpacing(12)

        lbl_in = QLabel("🎙️ Input:")
        lbl_in.setMinimumWidth(80)
        lbl_in.setObjectName("fieldLabel")
        inp_row.addWidget(lbl_in)

        self.input_device_combo = QComboBox()
        self.input_device_combo.setObjectName("inputDeviceCombo")
        inp_row.addWidget(self.input_device_combo, 1)
        self.input_device_combo.setVisible(False)

        self.input_device_btn = QPushButton("▾")
        self.input_device_btn.setObjectName("chevBtn")
        self.input_device_btn.setCursor(Qt.PointingHandCursor)
        self.input_device_btn.setFixedSize(34, 34)
        self.input_device_btn.clicked.connect(lambda: self._open_device_picker("input"))
        inp_row.addWidget(self.input_device_btn, 0)

        self.record_btn = QPushButton("● Start")
        self.record_btn.setObjectName("recordBtn")
        self.record_btn.clicked.connect(self.toggle_recording)
        self.record_btn.setMinimumWidth(120)
        inp_row.addWidget(self.record_btn)

        self.rec_timer_label = QLabel("00:00")
        self._set_monospace_timer_font(self.rec_timer_label)  # fixed-pitch digits for timers
        self._size_timer_label(self.rec_timer_label)

        self.rec_timer_label.setObjectName("recTimer")
        inp_row.addWidget(self.rec_timer_label)

        self.input_meter = QProgressBar(self.input_row_frame)
        self.input_meter.setRange(0, 100)
        self.input_meter.setTextVisible(False)
        self.input_meter.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self.input_meter.lower()
        self.input_meter.setStyleSheet("""
            QProgressBar { border:none; background:transparent; border-radius:9px; }
            QProgressBar::chunk { background:rgba(46,160,67,0.25); border-radius:9px; }
        """)

        self.input_caption = QLabel("", self.input_row_frame)
        self.input_caption.setObjectName("rowCaption")
        self.input_caption.hide()

        self.output_row_frame = QFrame()
        self.output_row_frame.setObjectName("outputRowFrame")
        self.output_row_frame.setProperty("stage", "idle")
        self.output_row_frame.setStyleSheet("""
            #outputRowFrame { background: #202e3a; border: 1px solid #304354; border-radius: 12px; }
        """)
        out_row = QHBoxLayout(self.output_row_frame)
        out_row.setContentsMargins(12, 10, 12, 10)
        out_row.setSpacing(12)

        lbl_out = QLabel("🔊 Output:")
        lbl_out.setMinimumWidth(80)
        lbl_out.setObjectName("fieldLabel")
        out_row.addWidget(lbl_out)

        self.output_device_combo = QComboBox()
        self.output_device_combo.setObjectName("outputDeviceCombo")
        out_row.addWidget(self.output_device_combo, 1)
        self.output_device_combo.setVisible(False)

        self.output_device_btn = QPushButton("▾")
        self.output_device_btn.setObjectName("chevBtn")
        self.output_device_btn.setCursor(Qt.PointingHandCursor)
        self.output_device_btn.setFixedSize(34, 34)
        self.output_device_btn.clicked.connect(lambda: self._open_device_picker("output"))
        out_row.addWidget(self.output_device_btn, 0)

        self.output_meter = QProgressBar(self.output_row_frame)
        self.output_meter.setRange(0, 100)
        self.output_meter.setTextVisible(False)
        self.output_meter.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self.output_meter.lower()
        self.output_meter.setStyleSheet("""
            QProgressBar { border:none; background:transparent; border-radius:9px; }
            QProgressBar::chunk { background:rgba(217,54,62,0.26); border-radius:9px; }
        """)
        self.output_caption = QLabel("", self.output_row_frame)
        self.output_caption.setObjectName("rowCaption")
        self.output_caption.hide()

        io_row = QHBoxLayout()
        io_row.setSpacing(12)
        io_row.addWidget(self.input_row_frame, 1)
        io_row.addWidget(self.output_row_frame, 1)
        dev_col.addLayout(io_row)

        def vsep():
            line = QFrame()
            line.setFrameShape(QFrame.VLine)
            line.setFrameShadow(QFrame.Plain)
            line.setStyleSheet("color:#2b4152;")
            return line

        top_row = QHBoxLayout(); top_row.setSpacing(10)
        self._top_row = top_row  # keep handle; widgets are moved later
        top_row.setContentsMargins(0, 0, 0, 0)

        src_lbl = QLabel("🌐 Source:"); src_lbl.setObjectName("fieldLabel"); top_row.addWidget(src_lbl)

        self.source_lang_input = QLineEdit("")
        self.target_lang_input = QLineEdit("")
        for w in (self.source_lang_input, self.target_lang_input):
            w.setMinimumWidth(0)
            w.setMaximumWidth(240)
            w.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        top_row.addWidget(self.source_lang_input, 0)

        self.swap_btn = QPushButton("⇄")
        self.swap_btn.setObjectName("swapBtn")
        self.swap_btn.setFixedWidth(44)
        self.swap_btn.setMinimumHeight(30)
        self.swap_btn.setToolTip("Swap Source ↔ Target (Ctrl+Enter / ⌘Enter)")
        self.swap_btn.clicked.connect(self.swap_languages)
        if sys.platform == "darwin": self.swap_btn.setShortcut("Meta+Return")
        else:                         self.swap_btn.setShortcut("Ctrl+Enter")
        top_row.addWidget(self.swap_btn)

        tgt_lbl = QLabel("🎯 Target:"); tgt_lbl.setObjectName("fieldLabel"); top_row.addWidget(tgt_lbl)
        top_row.addWidget(self.target_lang_input, 0)

        top_row.addWidget(vsep())

        mdl_lbl = QLabel("🤖 GPT Model:"); mdl_lbl.setObjectName("fieldLabel"); top_row.addWidget(mdl_lbl)
        self.gpt_model_input = QLineEdit("gpt-4.1-2025-04-14")
        self.gpt_model_input.setFixedWidth(220)
        top_row.addWidget(self.gpt_model_input)

        self.set_key_btn = QPushButton("🔑 Set API Key"); self.set_key_btn.clicked.connect(self._set_api_key)
        self.set_key_btn.setToolTip("Stored in memory only until you quit. Not saved to disk.")
        top_row.addWidget(self.set_key_btn, 0)

        _key_h = self.set_key_btn.sizeHint().height()
        _key_w = self.set_key_btn.sizeHint().width()

        self.stage_chip = QLabel("Idle")
        self._idle_logo = None
        for name in ("ccmi_logo.png", "logo.png"):
            p = resource_path(name)
            if os.path.exists(p):
                self._idle_logo = QPixmap(p)
                break
        self.stage_chip.setObjectName("stageChip")
        self.stage_chip.setProperty("state", "idle")
        self.stage_chip.setAlignment(Qt.AlignCenter)
        self.stage_chip.setMinimumHeight(_key_h)
        self.stage_chip.setMaximumHeight(_key_h)
        self.stage_chip.setMinimumWidth(_key_w)
        top_row.addWidget(self.stage_chip, 0)

        # Action timer: mirror record timer sizing/monospace for visual parity.
        timers = [w for w in self.findChildren(QLabel) if w.objectName() == "actionTimer"]
        if timers:
            self.action_timer_label = timers[0]
            for w in timers[1:]:
                try: w.setParent(None)
                except Exception: pass
                w.deleteLater()
        else:
            self.action_timer_label = QLabel("00:00", self)
            self.action_timer_label.setObjectName("actionTimer")

        self.action_timer_label.setFont(self.rec_timer_label.font())
        self.action_timer_label.setTextFormat(Qt.PlainText)
        self.action_timer_label.setWordWrap(False)
        self.action_timer_label.setAlignment(Qt.AlignCenter)
        self.action_timer_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self._set_monospace_timer_font(self.action_timer_label)
        self._size_timer_label(self.action_timer_label)

        top_row.addWidget(self.action_timer_label, 0)

        top_row.addWidget(vsep())
        top_row.addStretch(1)

        self.session_badge = QPushButton("Choose session…")
        self.session_badge.setObjectName("sessBadge")
        self.session_badge.setCursor(Qt.PointingHandCursor)
        self.session_badge.clicked.connect(self._enter_session_picker)
        self.session_badge.setMinimumHeight(_key_h)
        self.session_badge.setMaximumHeight(_key_h)
        self.session_badge.setMinimumWidth(_key_w)
        top_row.addWidget(self.session_badge, 0, Qt.AlignRight)
        self._sync_pill_sizes()

        self._refresh_api_key_button_style()

        top_row_wrap = QWidget()
        top_row_wrap.setLayout(top_row)

        self.input_row_frame.setMaximumWidth(180)
        self.output_row_frame.setMaximumWidth(180)

        io_row.addWidget(top_row_wrap, 2)

        io_row.setStretch(0, 0)
        io_row.setStretch(1, 0)
        io_row.setStretch(2, 2)

        dev_card.setLayout(dev_col)
        root.addWidget(dev_card, 0)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        setup_page = QWidget()
        self._setup_page = setup_page
        s = QVBoxLayout(setup_page); s.setSpacing(14)

        session_box = QGroupBox("Session")
        session_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.MinimumExpanding)
        self._session_box = session_box
        session_box.setStyleSheet(card_css + """
            QPushButton#cardToggle {
                background: transparent;      /* we custom-paint it */
                color: #eef6ff;
                border: none;                  /* border drawn in paintEvent */
                border-radius: 20px;
                padding: 26px;
                min-height: 380px;
                text-align: center;
            }

            QLabel#cardEmoji  { font-size: 56px; }
            QLabel#cardTitle  { font-size: 34px; font-weight: 900; letter-spacing:.2px; padding: 0 14px; }
            QLabel#cardSubtitle { font-size: 22px; font-weight: 800; color:#bcd6e7; margin-top: 4px; padding: 0 14px; }
            QLabel#cardDesc   { font-size: 20px; font-weight: 800; color:#eaf6ff; margin-top: 14px; padding: 0 18px; }

        """)

        session_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        sess_col = QVBoxLayout(session_box); sess_col.setSpacing(12)

        self.mode_one = QRadioButton("One-way");   self.mode_one.hide()
        self.mode_two = QRadioButton("Two-party"); self.mode_two.hide()
        self.chk_audience = QCheckBox("Audience present"); self.chk_audience.hide()
        
        self._session_group = QButtonGroup(self)
        self._session_group.setExclusive(True)
        self._session_group.addButton(self.mode_one)
        self._session_group.addButton(self.mode_two)

        self.session_stack_holder = QWidget()
        self.session_stack = QStackedLayout(self.session_stack_holder)
        self.session_stack_holder.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.session_stack.setContentsMargins(0, 0, 0, 0)
        cards_page = QWidget()
        cards_lay = QHBoxLayout(cards_page)
        cards_lay.setSpacing(16)
        cards_lay.setContentsMargins(12, 0, 12, 6)

        self.card_one = self._make_card(
            "🎯", "One-Way", "Source → Target",
            "Single speaker addressing an audience — perfect for announcements and lectures.",
            bg_path=img_one
        )
        self.card_two = self._make_card(
            "🤝", "Two-Party", "Both Source ↔ Target\n(no extra audience)",
            "Ideal for conversations and 1-to-1 calls — e.g., business meetings or negotiations.",
            bg_path=img_two
        )
        self.card_two_aud = self._make_card(
            "🎤", "Two-Party + Audience",
            "Both Source ↔ Target\n(audience present)",
            "Great when a conversation is open to third parties — panels, interviews, press Q&A.",
            bg_path=img_two_aud
        )

        CARD_MIN_W = 300
        for c in (self.card_one, self.card_two, self.card_two_aud):
            c.setMinimumWidth(CARD_MIN_W)
            c.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            cards_lay.addWidget(c, 1)
            cards_lay.setStretch(0, 1)
            cards_lay.setStretch(1, 1)
            cards_lay.setStretch(2, 1)

        self.session_stack.addWidget(cards_page)

        # No alignment here; allows horizontal expansion inside the box.
        sess_col.addWidget(self.session_stack_holder, 1)

        summary_page = QWidget(); sm = QHBoxLayout(summary_page); sm.setSpacing(10)
        self.session_summary = QLabel("Session: —"); self.session_summary.setObjectName("sessSummary")
        self.btn_change_session = QPushButton("Change session"); self.btn_change_session.setObjectName("changeBtn")
        self.btn_change_session.clicked.connect(self._enter_session_picker)
        sm.addWidget(self.session_summary, 1)
        sm.addWidget(self.btn_change_session, 0)

        self.session_stack.addWidget(summary_page)

        self.card_one.clicked.connect(lambda: self._select_session("one"))
        self.card_two.clicked.connect(lambda: self._select_session("two"))
        self.card_two_aud.clicked.connect(lambda: self._select_session("two_aud"))

        s.addWidget(session_box, 1)

        self.audience_box = QGroupBox("Audience")
        self.audience_box.setStyleSheet(card_css)
        af = QFormLayout(); af.setSpacing(8)

        self.lbl_aud_info   = QLabel("👥 Audience Info:")
        self.lbl_aud_expect = QLabel("🎯 Audience Expectation:")

        self.aud_persona  = QLineEdit()
        self.aud_guidance = QLineEdit()

        # Neutral defaults; overridden per session in _apply_mode_labels().
        self.aud_persona.setPlaceholderText("Audience (role / size)")
        self.aud_guidance.setPlaceholderText("What they need")

        af.addRow(self.lbl_aud_info,   self.aud_persona)
        af.addRow(self.lbl_aud_expect, self.aud_guidance)
        self.audience_box.setLayout(af)
        s.addWidget(self.audience_box, 0)

        presets_wrap = QWidget()
        self.presets_grid = QGridLayout(presets_wrap)
        self.presets_grid.setSpacing(14)

        preset_st = QGroupBox("Source → Target Translation Brief")
        preset_st.setStyleSheet(card_css)
        p1 = QFormLayout(); p1.setSpacing(8)
        self.st_brief   = QLineEdit(); self.st_brief.setPlaceholderText("one sentence on purpose")
        self.st_speaker = QLineEdit(); self.st_speaker.setPlaceholderText("role + flavor (e.g., ‘CEO, calm, concise’)")
        self.st_tone    = QLineEdit(); self.st_tone.setPlaceholderText("e.g., ‘neutral, formal(2)’")
        self.st_expect  = QLineEdit(); self.st_expect.setPlaceholderText("e.g., ‘faithful & fluent; keep style’")
        self.st_notes   = QTextEdit(); self.st_notes.setPlaceholderText("Notes / Constraints (anything else)")
        # Keep label refs so we can rename per mode.
        self.st_lbl_brief   = QLabel("📝 Purpose:")
        self.st_lbl_speaker = QLabel("👤 Source Speaker Info:")
        self.st_lbl_tone    = QLabel("🎛️ Tone:")
        self.st_lbl_expect  = QLabel("✅ Source Speaker Expectation:")
        self.st_lbl_notes   = QLabel("📌 Notes / Constraints:")
        p1.addRow(self.st_lbl_brief,   self.st_brief)
        p1.addRow(self.st_lbl_speaker, self.st_speaker)
        p1.addRow(self.st_lbl_tone,    self.st_tone)
        p1.addRow(self.st_lbl_expect,  self.st_expect)
        p1.addRow(self.st_lbl_notes,   self.st_notes)

        self.st_voice_select = QComboBox()
        self._populate_voice_combo(self.st_voice_select)
        self.st_test_tts_btn = QPushButton("🧪 Hear")
        self.st_test_tts_btn.setCursor(Qt.PointingHandCursor)
        self.st_test_tts_btn.clicked.connect(lambda: self._test_section_voice("st"))

        st_voice_wrap = QWidget()
        _st_vrow = QHBoxLayout(st_voice_wrap); _st_vrow.setContentsMargins(0,0,0,0); _st_vrow.setSpacing(8)
        _st_vrow.addWidget(self.st_voice_select); _st_vrow.addWidget(self.st_test_tts_btn, 0)
        
        self.btn_record_brief_st = QPushButton("🎤 Record Brief")
        self.btn_record_brief_st.setObjectName("briefBtn")
        self.btn_record_brief_st.setCursor(Qt.PointingHandCursor)
        try:
            hh = self.st_test_tts_btn.sizeHint().height()
        except Exception:
            hh = 36
        self.btn_record_brief_st.setMinimumHeight(hh)
        self.btn_record_brief_st.setMinimumWidth(120)
        self.btn_record_brief_st.clicked.connect(self.record_brief)
        _st_vrow.addWidget(self.btn_record_brief_st, 0)

        p1.addRow(QLabel("🗣️ TTS Voice:"), st_voice_wrap)
        
        preset_st.setLayout(p1)
        self._preset_st_box = preset_st

        preset_ts = QGroupBox("Target → Source Translation Brief")
        preset_ts.setStyleSheet(card_css)
        p2 = QFormLayout(); p2.setSpacing(8)
        self.ts_brief   = QLineEdit(); self.ts_brief.setPlaceholderText("one sentence on purpose (B → A)")
        self.ts_speaker = QLineEdit(); self.ts_speaker.setPlaceholderText("role + flavor (e.g., ‘Ops Manager, direct’)")
        self.ts_tone    = QLineEdit(); self.ts_tone.setPlaceholderText("e.g., ‘neutral, formal(2)’")
        self.ts_expect  = QLineEdit(); self.ts_expect.setPlaceholderText("e.g., ‘faithful & fluent; concise’")
        self.ts_notes   = QTextEdit(); self.ts_notes.setPlaceholderText("Notes / Constraints (anything else)")
        self.ts_lbl_brief   = QLabel("📝 Purpose:")
        self.ts_lbl_speaker = QLabel("👤 Target Speaker Info:")
        self.ts_lbl_tone    = QLabel("🎛️ Tone:")
        self.ts_lbl_expect  = QLabel("✅ Target Speaker Expectation:")
        self.ts_lbl_notes   = QLabel("📌 Notes / Constraints:")
        p2.addRow(self.ts_lbl_brief,   self.ts_brief)
        p2.addRow(self.ts_lbl_speaker, self.ts_speaker)
        p2.addRow(self.ts_lbl_tone,    self.ts_tone)
        p2.addRow(self.ts_lbl_expect,  self.ts_expect)
        p2.addRow(self.ts_lbl_notes,   self.ts_notes)

        self.ts_voice_select = QComboBox()
        self._populate_voice_combo(self.ts_voice_select)
        self.ts_test_tts_btn = QPushButton("🧪 Hear")
        self.ts_test_tts_btn.setCursor(Qt.PointingHandCursor)
        self.ts_test_tts_btn.clicked.connect(lambda: self._test_section_voice("ts"))

        ts_voice_wrap = QWidget()
        _ts_vrow = QHBoxLayout(ts_voice_wrap); _ts_vrow.setContentsMargins(0,0,0,0); _ts_vrow.setSpacing(8)
        _ts_vrow.addWidget(self.ts_voice_select); _ts_vrow.addWidget(self.ts_test_tts_btn, 0)
        
        self.btn_record_brief_ts = QPushButton("🎤 Record Brief")
        self.btn_record_brief_ts.setObjectName("briefBtn")
        self.btn_record_brief_ts.setCursor(Qt.PointingHandCursor)
        try:
            hh2 = self.ts_test_tts_btn.sizeHint().height()
        except Exception:
            hh2 = 36
        self.btn_record_brief_ts.setMinimumHeight(hh2)
        self.btn_record_brief_ts.setMinimumWidth(120)
        self.btn_record_brief_ts.clicked.connect(self.record_brief)
        _ts_vrow.addWidget(self.btn_record_brief_ts, 0)

        p2.addRow(QLabel("🗣️ TTS Voice:"), ts_voice_wrap)
        preset_ts.setLayout(p2)
        self._preset_ts_box = preset_ts

        self.presets_grid.addWidget(self._preset_st_box, 0, 0)
        self.presets_grid.addWidget(self._preset_ts_box, 0, 1)
        s.addWidget(presets_wrap, 1)

        self._i_session  = s.indexOf(session_box)
        self._i_audience = s.indexOf(self.audience_box)
        self._i_presets  = s.indexOf(presets_wrap)

        s.setStretch(self._i_session, 1)
        s.setStretch(self._i_audience, 0)
        s.setStretch(self._i_presets, 0)

        self.tabs.addTab(setup_page, "Setup")

        term_page = QWidget()
        t = QVBoxLayout(term_page); t.setSpacing(14)

        terms_box = QGroupBox("📚 Termlist")
        terms_box.setStyleSheet(card_css)
        tlay = QVBoxLayout()

        tbtns = QHBoxLayout()
        self.btn_import_terms = QPushButton("📥 Import Termlist (XLSX/CSV)")
        self.btn_import_terms.clicked.connect(self._import_termlist)
        self.btn_add_term = QPushButton("➕ Add Row")
        self.btn_add_term.clicked.connect(self._add_term_row)
        self.btn_del_term = QPushButton("🗑️ Delete Row(s)")
        self.btn_del_term.clicked.connect(self._delete_term_rows)
        self.btn_clear_terms = QPushButton("🧹 Clear")
        self.btn_clear_terms.clicked.connect(self._clear_terms)
        tbtns.addWidget(self.btn_import_terms)
        tbtns.addWidget(self.btn_add_term)
        tbtns.addWidget(self.btn_del_term)
        tbtns.addWidget(self.btn_clear_terms)
        tbtns.addStretch(1)
        tlay.addLayout(tbtns)

        self.termlist_table = QTableWidget(0, 2)
        self.termlist_table.setHorizontalHeaderLabels(["Source Term", "Target Term"])
        self.termlist_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.termlist_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.termlist_table.setAlternatingRowColors(True)
        self.termlist_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.termlist_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        tlay.addWidget(self.termlist_table, 1)

        terms_box.setLayout(tlay)
        t.addWidget(terms_box, 1)
        self.tabs.addTab(term_page, "Termlist")

        review_page = QWidget()
        self._review_page = review_page
        r = QVBoxLayout(review_page); r.setSpacing(12)

        seg_card = QGroupBox("🧾 Segments")
        seg_card.setStyleSheet(card_css)
        seg_col = QVBoxLayout()
        actions = QHBoxLayout()

        self.copy_last_btn = QPushButton("📋 Copy Last")
        self.copy_last_btn.clicked.connect(self._copy_last_translation)

        self.export_xlsx_btn = QPushButton("📤 Export XLSX")
        self.export_xlsx_btn.clicked.connect(self._export_xlsx)

        self.clear_btn = QPushButton("🧹 Clear")
        self.clear_btn.clicked.connect(self._clear_segments)

        actions.addWidget(self.copy_last_btn)
        actions.addWidget(self.export_xlsx_btn)
        actions.addWidget(self.clear_btn)
        actions.addStretch(1)
        seg_col.addLayout(actions)

        self.segments_table = QTableWidget(0, 2)
        self.segments_table.setHorizontalHeaderLabels(["Source", "Translation"])
        self.segments_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.segments_table.setShowGrid(True)
        self.segments_table.setWordWrap(True)
        self.segments_table.setEditTriggers(self.segments_table.NoEditTriggers)
        self.segments_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.segments_table.setAlternatingRowColors(True)
        seg_col.addWidget(self.segments_table)
        seg_card.setLayout(seg_col)
        r.addWidget(seg_card, 1)

        self.tabs.addTab(review_page, "Review")
        self.tabs.setTabText(self.tabs.indexOf(setup_page),  "Setup 🧰")
        self.tabs.setTabText(self.tabs.indexOf(term_page),   "Termlist 📚")
        self.tabs.setTabText(self.tabs.indexOf(review_page), "Review 🧾")

        self.input_row_frame.installEventFilter(self)
        self.output_row_frame.installEventFilter(self)

        self._populate_device_combos()
        self._update_device_badges()
        self.input_device_combo.currentIndexChanged.connect(self.on_input_device_changed)
        self.output_device_combo.currentIndexChanged.connect(self.on_output_device_changed)

        self._update_empty_state()
        self._add_card_shadows()

        self._set_status("idle")
        self._last_mode = "none"
        self._update_sections_from_toggles()

        root.removeWidget(dev_card)
        root.insertWidget(0, self.tabs, 1)
        root.insertWidget(1, dev_card, 0)
        
        # Reparent Idle+action timer+Session to tab-bar corner; then put Start+rec timer on top row.
        # Remove from old layouts before reparenting; sync sizes to avoid layout glitches.
        inp_row = self.input_row_frame.layout()
        try:
            inp_row.removeWidget(self.record_btn)
            inp_row.removeWidget(self.rec_timer_label)
        except Exception:
            pass

        try:
            self._top_row.removeWidget(self.stage_chip)
            self._top_row.removeWidget(self.action_timer_label)
            self._top_row.removeWidget(self.session_badge)
        except Exception:
            pass

        corner = QWidget(self.tabs)
        corner_lay = QHBoxLayout(corner)
        corner_lay.setContentsMargins(0, 0, 0, 0)
        corner_lay.setSpacing(12)
        corner_lay.addWidget(self.stage_chip)
        corner_lay.addWidget(self.action_timer_label)
        corner_lay.addWidget(self.session_badge)
        corner_lay.setAlignment(Qt.AlignVCenter | Qt.AlignRight)
        corner_lay.setContentsMargins(0, 0, 10, 0)
        corner_lay.insertSpacing(1, 8)
        corner_lay.insertSpacing(3, 8)

        self.tabs.setCornerWidget(corner, Qt.TopRightCorner)
        QTimer.singleShot(0, self._sync_corner_widget_heights)
        self._sync_corner_widget_heights()
        self._sync_pill_sizes()

        self._top_row.addWidget(vsep())
        self._top_row.addWidget(self.record_btn, 0)
        self._top_row.addWidget(self.rec_timer_label, 0)

        self._sync_pill_sizes()

    def _build_title_bar(self):
        """Custom frameless titlebar with Min and Close buttons."""
        self._title_bar = QFrame(self)
        self._title_bar.setObjectName("titleBar")
        self._title_bar.setFixedHeight(self._title_h)

        lay = QHBoxLayout(self._title_bar)
        lay.setContentsMargins(0, 0, 8, 0)
        lay.setSpacing(6)
        lay.addStretch(1)

        self._btn_min = QPushButton("–", self._title_bar)
        self._btn_min.setObjectName("btnMin")
        self._btn_min.setFixedSize(26, 22)
        self._btn_min.setToolTip("Minimize")
        self._btn_min.clicked.connect(self.showMinimized)
        lay.addWidget(self._btn_min, 0, Qt.AlignRight | Qt.AlignVCenter)

        self._btn_close = QPushButton("✕", self._title_bar)
        self._btn_close.setObjectName("btnClose")
        self._btn_close.setFixedSize(26, 22)
        self._btn_close.setToolTip("Close")
        self._btn_close.clicked.connect(self.close)
        lay.addWidget(self._btn_close, 0, Qt.AlignRight | Qt.AlignVCenter)

    def _sync_pill_sizes(self, min_w: int = 160):
        """Match Idle chip and Session badge to a common width/height.

        Uses Set API Key button size as baseline; ensures pills don't jitter.
        """
        key_h = max(self.set_key_btn.sizeHint().height(), 34)
        natural_idle  = self.stage_chip.sizeHint().width()
        natural_sess  = self.session_badge.sizeHint().width()
        base_w = max(self.set_key_btn.sizeHint().width(), min_w, natural_idle, natural_sess)

        for w in (self.stage_chip, self.session_badge):
            w.setMinimumWidth(base_w)
            w.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
            w.updateGeometry()

    def _set_api_key(self, on_success=None) -> bool:
        """Prompt for API key (session-only, not persisted); optional callback on success."""
        dlg = QDialog(self)
        dlg.setWindowTitle("OpenAI API Key")
        dlg.setModal(True)
        dlg.setWindowFlag(Qt.WindowContextHelpButtonHint, False)

        lay = QVBoxLayout(dlg)
        lay.setSpacing(10)

        title = QLabel("Enter your OpenAI API key")
        title.setStyleSheet("font-weight:800;")
        lay.addWidget(title)

        hint = QLabel("🔒 Your key is stored in memory only until you quit. "
                      "It is not saved to disk or printed to logs.")
        hint.setWordWrap(True)
        hint.setStyleSheet("color:#a6c3d9;")
        lay.addWidget(hint)

        edit = QLineEdit()
        edit.setEchoMode(QLineEdit.Password)
        edit.setPlaceholderText("sk-...")
        edit.setMinimumWidth(520)
        lay.addWidget(edit)

        show_cb = QCheckBox("Show key")
        def _toggle_show(v):
            edit.setEchoMode(QLineEdit.Normal if v else QLineEdit.Password)
        show_cb.toggled.connect(_toggle_show)
        lay.addWidget(show_cb)

        qa = QHBoxLayout()
        btn_paste = QPushButton("Paste")
        btn_paste.clicked.connect(lambda: edit.setText(QApplication.clipboard().text().strip()))
        btn_get = QPushButton("Get a key")
        btn_get.clicked.connect(lambda: QDesktopServices.openUrl(QUrl("https://platform.openai.com/account/api-keys")))
        qa.addStretch(1)
        qa.addWidget(btn_paste)
        qa.addWidget(btn_get)
        lay.addLayout(qa)

        err = QLabel("")
        err.setStyleSheet("color:#ff6b6b; font-weight:700;")
        err.hide()
        lay.addWidget(err)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(btns)

        def _handle_ok():
            k = (edit.text() or "")
            k = re.sub(r"\s+", "", k)
            k = k.replace("\u200b", "")

            if not (k.startswith("sk-") and len(k) >= 24):
                err.setText("That doesn’t look like an OpenAI API key (should start with “sk-”).")
                err.show()
                return

            set_api_key(k)
            self._refresh_api_key_button_style()
            self._show_toast("API key set for this session.")
            dlg.accept()
            if callable(on_success):
                QTimer.singleShot(0, on_success)

        btns.accepted.connect(_handle_ok)
        btns.rejected.connect(dlg.reject)

        ok = (dlg.exec_() == QDialog.Accepted)
        if ok and get_api_key():
            return True
        if not ok:
            self._show_toast("Cancelled.")
        elif not get_api_key():
            self._show_toast("No key entered.")
        return False

    def _require_api_key_then(self, fn) -> bool:
        """If key missing, open dialog and run `fn` via callback upon accept.

        Returns:
          True: key already present (no dialog shown).
          False: dialog shown; `fn` may run later.
        """
        if get_api_key():
            return True
        self._set_api_key(on_success=fn)
        return False

    def _toggle_max_restore(self):
        """Double-click titlebar toggles maximize/restore (frameless window)."""
        if self.isMaximized():
            self.showNormal()
        else:
            self.showMaximized()

    def _over_win_buttons(self, pos: QPoint) -> bool:
        """Return True if `pos` overlaps Min/Close; avoids intercepting their clicks."""
        if not hasattr(self, "_title_bar"):
            return False
        for b in (getattr(self, "_btn_min", None), getattr(self, "_btn_close", None)):
            if not b:
                continue
            gp = b.mapTo(self, QPoint(0, 0))
            if QRect(gp, b.size()).contains(pos):
                return True
        return False

    def mousePressEvent(self, e):
        """Start drag when pressing inside custom titlebar, excluding window buttons."""
        if e.button() == Qt.LeftButton and e.pos().y() <= self._title_h and not self._over_win_buttons(e.pos()):
            self._drag_pos = e.globalPos() - self.frameGeometry().topLeft()
            e.accept()
        super().mousePressEvent(e)

    def mouseMoveEvent(self, e):
        """Move window while dragging; disabled when maximized."""
        if self._drag_pos is not None and not self.isMaximized():
            self.move(e.globalPos() - self._drag_pos)
            e.accept()
        super().mouseMoveEvent(e)

    def mouseReleaseEvent(self, e):
        """End drag gesture."""
        self._drag_pos = None
        super().mouseReleaseEvent(e)

    def mouseDoubleClickEvent(self, e):
        """Double-click titlebar toggles maximize/restore; ignore clicks on buttons."""
        if e.button() == Qt.LeftButton and e.pos().y() <= self._title_h and not self._over_win_buttons(e.pos()):
            self._toggle_max_restore()
            e.accept()
        super().mouseDoubleClickEvent(e)

    def _add_term_row(self):
        """Append one empty (source, target) row."""
        row = self.termlist_table.rowCount()
        self.termlist_table.insertRow(row)
        for c in range(2):
            self.termlist_table.setItem(row, c, QTableWidgetItem(""))

    def _delete_term_rows(self):
        """Delete selected rows; toast when none selected."""
        if not self.termlist_table.selectionModel().hasSelection():
            self._show_toast("No rows selected.")
            return
        rows = sorted({idx.row() for idx in self.termlist_table.selectionModel().selectedRows()}, reverse=True)
        for r in rows:
            self.termlist_table.removeRow(r)
        self._show_toast(f"Deleted {len(rows)} row(s).")

    def _clear_terms(self):
        """Remove all term rows."""
        self.termlist_table.setRowCount(0)
        self._show_toast("Termlist cleared.")

    def _import_termlist(self):
        """Load terms from XLSX (openpyxl) or CSV; trims cells; skips empty rows."""
        path, _ = QFileDialog.getOpenFileName(self, "Import Termlist", "",
                                              "Excel (*.xlsx);;CSV (*.csv)")
        if not path:
            return
        try:
            rows = []
            if path.lower().endswith(".xlsx"):
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    self._show_toast("openpyxl not installed; import a CSV or install openpyxl.")
                    return
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                for r in ws.iter_rows(values_only=True):
                    if r is None: continue
                    s = (str(r[0]).strip() if len(r) > 0 and r[0] is not None else "")
                    t = (str(r[1]).strip() if len(r) > 1 and r[1] is not None else "")
                    if s or t:
                        rows.append((s, t))
            else:
                with open(path, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if not row:
                            continue
                        s = (row[0].strip() if len(row) > 0 and row[0] is not None else "")
                        t = (row[1].strip() if len(row) > 1 and row[1] is not None else "")
                        if s or t:
                            rows.append((s, t))
            self.termlist_table.setRowCount(0)
            for sterm, tterm in rows:
                r = self.termlist_table.rowCount()
                self.termlist_table.insertRow(r)
                self.termlist_table.setItem(r, 0, QTableWidgetItem(sterm))
                self.termlist_table.setItem(r, 1, QTableWidgetItem(tterm))
            self._show_toast(f"Loaded {len(rows)} term pairs.")
        except Exception as e:
            self._show_toast(f"Import error: {e}")

    def _build_term_lines(self, reverse: bool) -> str:
        """Build 'src = tgt' lines from table; swap when reverse=True."""
        lines = []
        for r in range(self.termlist_table.rowCount()):
            s_it = self.termlist_table.item(r, 0)
            t_it = self.termlist_table.item(r, 1)
            s = "" if s_it is None else s_it.text().strip()
            t = "" if t_it is None else t_it.text().strip()
            if not s and not t:
                continue
            if reverse:
                s, t = t, s
            if s and t:
                lines.append(f"{s} = {t}")
        return "\n".join(lines)

    def _add_card_shadows(self):
        """Attach drop shadows to all QGroupBox cards."""
        for gb in self.findChildren(QGroupBox):
            eff = QGraphicsDropShadowEffect(self)
            eff.setBlurRadius(18)
            eff.setXOffset(0)
            eff.setYOffset(6)
            eff.setColor(QColor(0, 0, 0, 140))
            gb.setGraphicsEffect(eff)

    def resizeEvent(self, e):
        """Keep overlays and custom titlebar aligned on resize."""
        super().resizeEvent(e)
        if hasattr(self, "toast") and self.toast:
            self.toast.move(max(0, self.width() - self.toast.width() - 24), 18)
        if hasattr(self, "help_overlay") and self.help_overlay:
            self.help_overlay.setGeometry(0, 0, self.width(), self.height())
        if hasattr(self, "_title_bar"):
            self._title_bar.setGeometry(0, 0, self.width(), self._title_h)
            self._title_bar.raise_()
        
        self._sync_corner_widget_heights()
    
    def fit_to_work_area(self):
        """Set fixed size to available screen and center the window."""
        avail = QApplication.primaryScreen().availableGeometry()
        W, H = 1920, 1080
        w = min(W, avail.width())
        h = min(H, avail.height())
        self.setFixedSize(w, h)
        self.move(avail.center() - self.rect().center())

    def _on_accent_changed(self, idx):
        """Map accent selector index → theme key and reapply theme."""
        mapping = {0: "emerald", 1: "cyan", 2: "violet", 3: "amber"}
        self._accent = mapping.get(idx, "emerald")
        self.apply_dark_theme(self._accent, self._theme_variant)

    def _set_status(self, state: str):
        """Update status pill and blink while recording (≈600 ms pulse)."""
        if not hasattr(self, "status_pill"):
            return
        self.status_pill.setText("Recording" if state == "recording" else "Idle")
        self.status_pill.setProperty("state", "recording" if state == "recording" else "idle")
        self.status_pill.setProperty("pulse", "off")
        self.status_pill.style().unpolish(self.status_pill)
        self.status_pill.style().polish(self.status_pill)

        if self._status_timer:
            self._status_timer.stop()
            self._status_timer.deleteLater()
            self._status_timer = None

        if state == "recording":
            self._status_timer = QTimer(self)
            self._status_timer.timeout.connect(self._toggle_pulse)
            self._status_timer.start(600)

    def _toggle_pulse(self):
        """Flip 'pulse' property to drive CSS animation."""
        cur = self.status_pill.property("pulse")
        self.status_pill.setProperty("pulse", "off" if cur == "on" else "on")
        self.status_pill.style().unpolish(self.status_pill)
        self.status_pill.style().polish(self.status_pill)

    def _start_rec_timer(self):
        """Start mm:ss timer; ticks every 1s."""
        self._record_ms = 0
        if self._record_timer:
            self._record_timer.stop()
        self._record_timer = QTimer(self)
        self._record_timer.timeout.connect(self._tick_rec)
        self._record_timer.start(1000)
        self.rec_timer_label.setText("00:00")

    def _stop_rec_timer(self):
        """Stop and reset the record timer display."""
        if self._record_timer:
            self._record_timer.stop()
        self.rec_timer_label.setText("00:00")

    def _tick_rec(self):
        """Increment elapsed time and update mm:ss label."""
        self._record_ms += 1000
        s = self._record_ms // 1000
        m = s // 60
        s = s % 60
        self.rec_timer_label.setText(f"{m:02d}:{s:02d}")

    def closeEvent(self, event):
        """Clean shutdown: stop threads, release audio, clear API key."""
        self.stop_volume_monitor()
        if self.recording_thread and self.recording_thread.isRunning():
            self.recording_thread.stop()
            self.recording_thread.wait()
        for t in (
            self.transcription_thread,
            self.translation_thread,
            self.tts_thread,
            self.brief_recording_thread,
            self.brief_transcription_thread,
            self.brief_filling_thread,
        ):
            try:
                if t and t.isRunning():
                    t.wait(5000)
            except Exception:
                pass
        set_api_key("")
        event.accept()

    def swap_languages(self):
        """Swap Source/Target fields and flip reverse_mode so voices follow direction."""
        self._update_stage_chip("idle", "Idle")
        src = self.source_lang_input.text()
        tgt = self.target_lang_input.text()
        self.source_lang_input.setText(tgt)
        self.target_lang_input.setText(src)

        self.reverse_mode = not self.reverse_mode

        self.append_console(
            f"Swapped languages: Source='{tgt}', Target='{src}' "
            f"(direction={'TS' if self.reverse_mode else 'ST'})"
        )

    def test_tts(self):
        """Quick voice check using a sample sentence."""
        sample = "Translate the purpose, not the words."
        self.append_console("Testing TTS…")
        self.start_tts_stage(sample)

    def apply_dark_theme(self, accent="emerald", variant=None):
        """Apply app-wide dark stylesheet with accent variants.

        Chips share Set-API-Key height for visual parity. After font changes,
        recompute monospace timer metrics to avoid width jitter/truncation.
        """
        p = {
            "CANVAS": "#182632",
            "PANEL": "#1f2f3b",
            "PANEL_ALT": "#223543",
            "BORDER": "#304354",
            "TEXT": "#eef6ff",
            "MUTED": "#c7d6e3",
            "TABLE_ODD": "#203241",
            "TABLE_EVEN": "#1d2e3b",
            "SELECT_BG": "#29465a",
        }
        accents = {
            "emerald": {"ACCENT": "#2ee59d", "ACCENT_BG": "#133b2d", "ACCENT_RING": "#35e0a7"},
            "cyan":    {"ACCENT": "#39c2ff", "ACCENT_BG": "#0f3346", "ACCENT_RING": "#39c2ff"},
            "violet":  {"ACCENT": "#b896ff", "ACCENT_BG": "#261a3f", "ACCENT_RING": "#b896ff"},
            "amber":   {"ACCENT": "#ffc36b", "ACCENT_BG": "#3c2a10", "ACCENT_RING": "#ffc36b"},
        }
        ac = accents.get(accent, accents["emerald"])
        
        # Use Set API Key's size as the standard chip height
        try:
            btn_h = max(self.set_key_btn.sizeHint().height(), 34)
        except Exception:
            btn_h = 36


        style = f"""
            QWidget {{
                background: {p['CANVAS']};
                color: {p['TEXT']};
                font-family: Inter, "Segoe UI", "Segoe UI Emoji", "Noto Color Emoji", "Apple Color Emoji", system-ui, Arial, sans-serif;
                font-size: 20px;
            }}

            QLabel {{ background: transparent; color: {p['TEXT']}; padding: 0; border: none; }}
            #fieldLabel {{ color: {ac['ACCENT']}; font-weight: 800; font-size: 17px; }}
            #recTimer  {{ color: {p['MUTED']}; font-weight: 700; }}
            #actionTimer {{
                color: #c7d6e3;
                font-weight: 700;
                padding: 0 6px;      /* extra comfort, pairs with width pad above */
            }}

            QLineEdit, QTextEdit, QTextBrowser, QComboBox, QTableWidget, QProgressBar {{
                background: {p['PANEL']}; color: {p['TEXT']};
                border: 1px solid {p['BORDER']}; border-radius: 10px; padding: 9px 12px;
                selection-background-color: {p['SELECT_BG']}; selection-color: {p['TEXT']};
            }}
            QLineEdit:hover, QTextEdit:hover, QComboBox:hover, QTableWidget:hover {{ background: {p['PANEL_ALT']}; }}
            QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QTableWidget:focus {{
                border: 2px solid {ac['ACCENT_RING']};
            }}

            QComboBox QAbstractItemView {{
                background: {p['PANEL']}; color: {p['TEXT']}; border: 1px solid {p['BORDER']};
                selection-background-color: {p['SELECT_BG']};
            }}

            QPushButton {{
                font-size: 16px; padding: 10px 18px; border-radius: 12px;
                background: #21353f; border: 1px solid {p['BORDER']}; color: {p['TEXT']}; font-weight: 700;
            }}
            QPushButton:hover  {{ background: #253b46; }}
            QPushButton:pressed{{ background: #1c2e37; }}
            #recordBtn {{ background: {ac['ACCENT_BG']}; border: 2px solid {ac['ACCENT_RING']}; color: #f3fbf8; }}
            #recordBtn[recording="true"] {{ background: #2b0e12; border-color:#d43a47; color:#fff2f2; }}
            #dangerBtn {{ background:#1a0e11; border:1px solid #d43a47; color:#ffecec; }}
            #dangerBtn:hover {{ background:#2b0e12; }}

            #swapBtn {{ font-size: 18px; padding: 0; min-width: 44px; }}

            /* === [PILLS — unified with Set API Key] ================================== */
            #sessBadge, #stageChip {{
                background: #203241;
                border: 1px solid #2b4152;
                color: #dbe8f5;
                font-weight: 700;
                font-size: 16px;              /* match Set API Key */
                padding: 8px 12px;
                border-radius: 10px;
            }}
            
            #devBadge {{
                background: #203241;
                border: 1px solid #2b4152;
                color: #dbe8f5;
                font-weight: 700;
                font-size: 14px;
                padding: 8px 10px;
                border-radius: 10px;
            }}
            #devBadge:hover {{ background: #253a49; }}
            
            
            #sessBadge:hover {{ background: #253a49; }}

            /* === [SESSION BADGE VARIANTS — paste below] ===================== */
            /* One-way  → cyan (no emerald/violet/red) */
            #sessBadge[mode="one"] {{
                background:#0f3346;   /* cyan ACCENT_BG */
                border-color:#39c2ff; /* cyan ring */
                color:#e6f6ff;
            }}
            /* Two-party → amber */
            #sessBadge[mode="two"] {{
                background:#3c2a10;   /* amber ACCENT_BG */
                border-color:#ffc36b; /* amber ring */
                color:#fff3df;
            }}
            /* Two-party + Audience → deep blue (coherent with slate/blue UI) */
            #sessBadge[mode="two_aud"] {{
                background:#1b2940;   /* deep blue */
                border-color:#6aa8ff; /* soft blue ring */
                color:#e6f0ff;
            }}


            /* Center QLabel text for Idle chip */
            #stageChip {{ qproperty-alignment: AlignCenter; }}

            /* Brief button lives beside Hear; turn red while recording */
            #briefBtn {{ font-weight: 800; }}
            #briefBtn[recording="true"] {{
                background:#2b0e12; border-color:#d43a47; color:#fff2f2;
            }}

            /* stage states */
            #stageChip[state="idle"]        {{ background:#211a33; border-color:#5f4b8b; color:#efeaff; }}
            #stageChip[state="idle_alt"]    {{ background:#261a3f; border-color:#8e79c9; color:#efeaff; }}
            #stageChip[state="busy"]        {{ background:#2a2147; border-color:#b896ff; color:#f2ebff; }}
            #stageChip[pulse="on"][state="busy"] {{ background:#352a5d; border-color:#c9b3ff; }}
            #stageChip[state="play"]        {{ background:#1c3b2f; border-color:#2ee59d; }}
            #stageChip[pulse="on"][state="play"] {{ background:#204e3d; border-color:#59f0bb; }}
            #stageChip[state="ok"]          {{ background:#1e2f39; border-color:#35e0a7; }}
            /* ======================================================================== */

            /* Idle — deep aubergine */
            #stageChip[state="idle"] {{
                background:#211a33;
                border-color:#5f4b8b;
                color:#efeaff;
            }}

            /* Optional alt idle (slightly brighter violet) */
            #stageChip[state="idle_alt"] {{
                background:#261a3f;
                border-color:#8e79c9;
                color:#efeaff;
            }}

            /* Busy (live waiting) — violet ink + pulse */
            #stageChip[state="busy"] {{
                background:#2a2147;
                border-color:#b896ff;
                color:#f2ebff;
            }}
            #stageChip[pulse="on"][state="busy"] {{
                background:#352a5d;
                border-color:#c9b3ff;
            }}

            /* Speaking & Done keep emerald cues */
            #stageChip[state="play"] {{
                background:#1c3b2f;
                border-color:#2ee59d;
            }}
            #stageChip[pulse="on"][state="play"] {{
                background:#204e3d;
                border-color:#59f0bb;
            }}
            #stageChip[state="ok"] {{
                background:#1e2f39;
                border-color:#35e0a7;
            }}

            /* IO row tints by stage */
            #inputRowFrame[stage="idle"]   {{ background:#211a3f; border-color:#5f4b8b; }}  /* deep aubergine */
            #inputRowFrame[stage="busy"]   {{ background:#2a2147; border-color:#b896ff; }}  /* waiting = violet */
            #inputRowFrame[stage="play"]   {{ background:#15352b; border-color:#2ee59d; }}  /* speaking = emerald */

            #outputRowFrame[stage="idle"]  {{ background:#211a3f; border-color:#5f4b8b; }}
            #outputRowFrame[stage="busy"]  {{ background:#2a2147; border-color:#b896ff; }}
            #outputRowFrame[stage="play"]  {{ background:#15352b; border-color:#2ee59d; }}

            /* (Optional) top card reacts too */
            #devCard[stage="busy"]         {{ background:#231b38; border-color:#8e79c9; }}
            #devCard[stage="play"]         {{ background:#12372b; border-color:#2ee59d; }}


            #chevBtn {{
                background: #203241;
                border: 1px solid #2b4152;
                color: #dbe8f5;
                font-weight: 900;
                font-size: 18px;
                padding: 0 8px;
                border-radius: 10px;
                min-width: 30px; max-width: 34px; min-height: 30px;
            }}
            #chevBtn:hover {{ background: #253a49; }}

            /* Tabs (fix white background by making the bar transparent) */
            QTabWidget::pane {{
                border: 1px solid {p['BORDER']};
                border-radius: 12px;
                top: 8px;
                background: {p['PANEL']};
            }}
            QTabWidget::tab-bar {{ background: transparent; }}
            QTabBar {{ background: transparent; }}
            QTabBar::tab {{
                background: {p['PANEL_ALT']};
                color: {p['TEXT']};
                padding: 9px 14px;
                margin-right: 6px;
                border: 1px solid {p['BORDER']};
                border-bottom: none;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
            }}
            QTabBar::tab:selected {{ background:#263a49; border-color:{ac['ACCENT_RING']}; }}
            QTabBar::tab:hover    {{ background:#243643; }}

            /* Tables */
            QHeaderView::section {{
                background: #223545; color: #e6fff5; font-size: 16px; font-weight: 800;
                border: none; border-bottom: 1px solid {p['BORDER']}; padding: 10px 12px; letter-spacing: .2px;
            }}
            QTableWidget {{
                background: {p['TABLE_EVEN']};
                gridline-color: {p['BORDER']};
                alternate-background-color: {p['TABLE_ODD']};
            }}
            QTableWidget::item:selected {{ background: #0f6a4d; color: #ffffff; }}

            /* Scrollbars */
            QScrollBar:vertical, QScrollBar:horizontal {{
                background: transparent; border: none; margin: 6px; width: 10px; height:10px;
            }}
            QScrollBar::handle:vertical, QScrollBar::handle:horizontal {{
                background: #355062; min-height: 24px; border-radius: 6px;
            }}
            QScrollBar::handle:vertical:hover, QScrollBar::handle:horizontal:hover {{ background: #3e5c71; }}

            #toast {{ background: rgba(26,38,48,0.94); border: 1px solid {p['BORDER']}; border-radius: 10px; padding: 10px 14px; color: {p['TEXT']}; }}
            #helpOverlay {{ background: rgba(22,32,41,0.92); border: 1px solid {p['BORDER']}; border-radius: 12px; color:{p['TEXT']}; }}
        """
        
        emerald_hover_bg = {
            "emerald": "#133b2d",
            "cyan":    "#0f3346",
            "violet":  "#261a3f",
            "amber":   "#3c2a10",
        }.get(accent, "#133b2d")

        style += f"""
            #titleBar {{
                background: transparent;     /* no white strip */
            }}
            QPushButton#btnMin, QPushButton#btnClose {{
                border: 1px solid transparent;
                border-radius: 6px;
                padding: 0;
                font-size: 14px;
                min-width: 26px;
                min-height: 22px;
                background: transparent;
                color: {p['TEXT']};
            }}
            /* Minimize → emerald on hover */
            QPushButton#btnMin:hover {{
                background: {emerald_hover_bg};
                border-color: {ac['ACCENT_RING']};
            }}
            /* Close → red on hover */
            QPushButton#btnClose:hover {{
                background: #2b0e12;
                border-color: #d43a47;
            }}
        """

        self.setStyleSheet(style)
        # Qt quirk: fonts affect label width; recompute fixed-width timers post-style.
        if hasattr(self, "action_timer_label"):
            self._set_monospace_timer_font(self.action_timer_label)
            self._size_timer_label(self.action_timer_label)
        if hasattr(self, "rec_timer_label"):
            self._set_monospace_timer_font(self.rec_timer_label)
            self._size_timer_label(self.rec_timer_label)

        self._sync_pill_sizes()

    def _start_action_timer(self):
        """Start mm:ss action timer; no-op if already running."""
        if getattr(self, "_action_timer", None) and self._action_timer.isActive():
            return
        self._action_ms = 0
        if self._action_timer:
            self._action_timer.stop()
        self._action_timer = QTimer(self)
        self._action_timer.timeout.connect(self._tick_action)
        self._action_timer.start(1000)
        if hasattr(self, "action_timer_label"):
            self.action_timer_label.setText("00:00")

    def _stop_action_timer(self):
        """Stop and reset the action timer."""
        if getattr(self, "_action_timer", None):
            self._action_timer.stop()
        if hasattr(self, "action_timer_label"):
            self.action_timer_label.setText("00:00")

    # === [TIMER FONT + SIZING HELPERS] ===========================================
    def _set_monospace_timer_font(self, lbl: QLabel):
        """Use fixed-pitch digits; preserve size; avoid Qt pointSize=-1 warnings."""
        if not lbl:
            return
        try:
            base = lbl.font()
            fixed = QFontDatabase.systemFont(QFontDatabase.FixedFont)
            fixed.setStyleHint(QFont.Monospace)
            fixed.setFixedPitch(True)
            fixed.setKerning(False)

            ps = base.pointSize()
            if ps > 0:
                fixed.setPointSize(ps)
            else:
                px = base.pixelSize()
                if px > 0:
                    fixed.setPixelSize(px)
                else:
                    fixed.setPointSize(13)

            fixed.setBold(True)
            lbl.setFont(fixed)

            fm = QFontMetrics(fixed)
            lbl.setMinimumHeight(max(lbl.minimumHeight(), fm.height() + 2))  # prevent vertical clipping
        except Exception:
            pass

    def _size_timer_label(self, lbl: QLabel):
        """Fix label width for widest '88:88' to prevent jitter; add padding."""
        if not lbl:
            return
        fm = QFontMetrics(lbl.font())
        advance = getattr(fm, "horizontalAdvance", fm.width)
        text_w = max(advance("88:88"), fm.boundingRect("88:88").width())

        pad_each = max(10, fm.averageCharWidth())
        css_lr = 12  # from #actionTimer { padding: 0 6px; }
        safety = 2

        total = text_w + (2 * pad_each) + css_lr + safety
        lbl.setFixedWidth(total)
        lbl.setMinimumWidth(total)
        lbl.setWordWrap(False)
        lbl.setMinimumHeight(max(lbl.minimumHeight(), fm.height() + 4))

    def _tick_action(self):
        """Increment action timer and update mm:ss label."""
        self._action_ms += 1000
        s = self._action_ms // 1000
        m = s // 60
        s = s % 60
        if hasattr(self, "action_timer_label"):
            self.action_timer_label.setText(f"{m:02d}:{s:02d}")

    def _sync_corner_widget_heights(self):
        """Match pills to the tab bar height; scale idle logo to pill height."""
        if not hasattr(self, "tabs") or not hasattr(self, "stage_chip"):
            return
        bar_h = self.tabs.tabBar().sizeHint().height()
        gap = 6
        pill_h = max(24, bar_h - gap)
        for w in (self.stage_chip, self.session_badge):
            w.setFixedHeight(pill_h)
        if hasattr(self, "action_timer_label"):
            fm = QFontMetrics(self.action_timer_label.font())
            self.action_timer_label.setFixedHeight(max(fm.height() + 2, pill_h - 2))

        if (getattr(self, "_idle_logo", None)
                and not self._idle_logo.isNull()
                and hasattr(self, "stage_chip")
                and self.stage_chip.property("state") == "idle"):
            self.stage_chip.setPixmap(self._idle_logo.scaledToHeight(max(16, pill_h - 6), Qt.SmoothTransformation))

    # -------------------------------------------------------------------------
    # Volume Monitoring
    # -------------------------------------------------------------------------
    def start_volume_monitor(self):
        """Start RMS→% volume monitor thread."""
        if not self.volume_monitor_thread:
            self.volume_monitor_thread = VolumeMonitorThread(self.fixed_input_device)
            self.volume_monitor_thread.volume_signal.connect(self.update_volume_bar)
            self.volume_monitor_thread.start()

    def stop_volume_monitor(self):
        """Stop and clear the volume monitor thread."""
        if self.volume_monitor_thread:
            self.volume_monitor_thread.stop()
            self.volume_monitor_thread = None

    def update_volume_bar(self, vol):
        """Update input meter with latest volume %."""
        if hasattr(self, "input_meter"):
            self.input_meter.setValue(vol)

    def _detect_default_devices(self):
        """Pick defaults via sounddevice; fall back to first usable in/out; raise if none."""
        try:
            d = sd.default.device
            if isinstance(d, (list, tuple)) and len(d) >= 2:
                di = None if d[0] is None else int(d[0])
                do = None if d[1] is None else int(d[1])
            else:
                di = do = None
        except Exception:
            di = do = None

        if di is None or do is None:
            try:
                ha = sd.query_hostapis(sd.default.hostapi)
                di = ha.get("default_input_device", di)
                do = ha.get("default_output_device", do)
            except Exception:
                pass

        devices = sd.query_devices()
        if di is None:
            for i, dev in enumerate(devices):
                if dev["max_input_channels"] > 0:
                    di = i
                    break
        if do is None:
            for i, dev in enumerate(devices):
                if dev["max_output_channels"] > 0:
                    do = i
                    break

        if di is None or do is None:
            raise RuntimeError("No usable audio devices were found.")
        return di, do

    def _list_input_devices(self):
        """Return [(id, label)] for devices with input channels."""
        return [(i, f"{dev['name']} (ID: {i})")
                for i, dev in enumerate(sd.query_devices())
                if dev["max_input_channels"] > 0]

    def _list_output_devices(self):
        """Return [(id, label)] for devices with output channels."""
        return [(i, f"{dev['name']} (ID: {i})")
                for i, dev in enumerate(sd.query_devices())
                if dev["max_output_channels"] > 0]

    def _is_recording(self):
        """True if the recording thread is active."""
        return self.recording_thread is not None and self.recording_thread.isRunning()

    def _restart_volume_monitor(self):
        """Restart monitor after device change."""
        self.stop_volume_monitor()
        self.start_volume_monitor()

    def on_input_device_changed(self, _idx):
        """Apply selected input device; block while recording."""
        if self._is_recording():
            self._populate_device_combos()
            self.append_console("Cannot change input device while recording.")
            return
        dev_id = self.input_device_combo.currentData()
        if isinstance(dev_id, int):
            self.fixed_input_device = dev_id
            self.append_console(f"Input device set to ID {dev_id}.")
            self._restart_volume_monitor()
        self._update_device_badges()    
        
    def on_output_device_changed(self, _idx):
        """Apply selected output device."""
        dev_id = self.output_device_combo.currentData()
        if isinstance(dev_id, int):
            self.fixed_output_device = dev_id
            self.append_console(f"Output device set to ID {dev_id}.")
        self._update_device_badges()

    def _populate_device_combos(self):
        """Fill device combos and restore current selections without signals."""
        self.input_device_combo.blockSignals(True)
        self.output_device_combo.blockSignals(True)

        self.input_device_combo.clear()
        for i, name in self._list_input_devices():
            self.input_device_combo.addItem(name, i)
        idx = self.input_device_combo.findData(self.fixed_input_device)
        if idx >= 0: self.input_device_combo.setCurrentIndex(idx)

        self.output_device_combo.clear()
        for i, name in self._list_output_devices():
            self.output_device_combo.addItem(name, i)
        idx = self.output_device_combo.findData(self.fixed_output_device)
        if idx >= 0: self.output_device_combo.setCurrentIndex(idx)

        self.input_device_combo.blockSignals(False)
        self.output_device_combo.blockSignals(False)
        
    # === [DEVICE BADGES + POPUP PICKER] ==========================================
    def _update_device_badges(self):
        """Update device pill tooltips with device name and ID."""
        try:
            devs = sd.query_devices()
            if hasattr(self, "input_device_btn"):
                name = devs[self.fixed_input_device]["name"] if 0 <= self.fixed_input_device < len(devs) else "?"
                self.input_device_btn.setText("▾")
                self.input_device_btn.setToolTip(f"{name}  (ID: {self.fixed_input_device})")
            if hasattr(self, "output_device_btn"):
                name = devs[self.fixed_output_device]["name"] if 0 <= self.fixed_output_device < len(devs) else "?"
                self.output_device_btn.setText("▾")
                self.output_device_btn.setToolTip(f"{name}  (ID: {self.fixed_output_device})")
        except Exception:
            pass

    def _open_device_picker(self, kind: str):
        """Frameless, translucent picker with live filter; double-click selects.

        Qt quirk: use WA_TranslucentBackground + inner chrome frame to avoid
        artifacts. Current device floats to top; selection updates hidden combos.
        """
        if kind == "input" and self._is_recording():
            self._show_toast("Stop recording to change the input device.")
            return

        dlg = QDialog(self)
        dlg.setModal(True)
        dlg.setWindowFlag(Qt.WindowContextHelpButtonHint, False)
        dlg.setWindowFlag(Qt.FramelessWindowHint, True)
        dlg.setAttribute(Qt.WA_TranslucentBackground, True)
        dlg.setMinimumWidth(620)

        root = QVBoxLayout(dlg)
        root.setContentsMargins(6, 6, 6, 6)

        chrome = QFrame(dlg)
        chrome.setObjectName("pickerChrome")
        chrome.setStyleSheet("""
            #pickerChrome {
                background: #1f2f3b;
                border: 1px solid #304354;
                border-radius: 12px;
            }
        """)
        root.addWidget(chrome)

        lay = QVBoxLayout(chrome)
        lay.setSpacing(10)
        lay.setContentsMargins(14, 14, 14, 14)

        hdr = QHBoxLayout()
        title = QLabel("Select a device")
        title.setStyleSheet("font-weight:800;")
        btn_close = QPushButton("✕")
        btn_close.setObjectName("btnClose")
        btn_close.setFixedSize(26, 22)
        btn_close.clicked.connect(dlg.reject)
        hdr.addWidget(title)
        hdr.addStretch(1)
        hdr.addWidget(btn_close)
        lay.addLayout(hdr)

        hint = QLabel("🔎 Type to filter • Double-click to select • Current device is pre-selected.")
        hint.setWordWrap(True)
        hint.setStyleSheet("color:#a6c3d9;")
        lay.addWidget(hint)

        search = QLineEdit()
        search.setPlaceholderText("Filter by name or ID…")
        lay.addWidget(search)

        lst = QListWidget(dlg)
        lst.setStyleSheet("""
            QListWidget {
                background: #1b2a36;
                color: #eef6ff;
                border: 1px solid #304354;
                border-radius: 8px;
                alternate-background-color: #203241;
                outline: none;
            }
            QListWidget::item {
                background: transparent;
                padding: 8px 10px;
            }
            QListWidget::item:alternate { background: #1d2e3b; }
            QListWidget::item:selected  { background: #29465a; color: #ffffff; }
            QListWidget::item:hover     { background: #243643; }
        """)
        lst.setAlternatingRowColors(True)

        pal = lst.palette()
        pal.setColor(QPalette.Base, QColor("#1b2a36"))
        pal.setColor(QPalette.AlternateBase, QColor("#1d2e3b"))
        pal.setColor(QPalette.Text, QColor("#eef6ff"))
        pal.setColor(QPalette.Highlight, QColor("#29465a"))
        pal.setColor(QPalette.HighlightedText, QColor("#ffffff"))
        lst.setPalette(pal)

        lst.setSelectionMode(QAbstractItemView.SingleSelection)
        lst.setUniformItemSizes(True)
        lst.setMinimumHeight(360)
        lay.addWidget(lst, 1)

        def populate():
            lst.clear()
            devices = sd.query_devices()

            if kind == "input":
                items = [(i, d) for i, d in enumerate(devices) if d["max_input_channels"] > 0]
                cur_id = self.fixed_input_device
                icon = "🎙️"
            else:
                items = [(i, d) for i, d in enumerate(devices) if d["max_output_channels"] > 0]
                cur_id = self.fixed_output_device
                icon = "🔊"

            # Current device first, then ID order.
            items.sort(key=lambda pair: (pair[0] != cur_id, pair[0]))

            sel_row = 0
            for row, (dev_id, d) in enumerate(items):
                tick = "✅ " if dev_id == cur_id else ""
                text = f"{icon}  {tick}{d['name']}   —   ID {dev_id}"
                it = QListWidgetItem(text, lst)
                it.setData(Qt.UserRole, dev_id)
                if dev_id == cur_id:
                    sel_row = row

            if lst.count():
                lst.setCurrentRow(sel_row)
                lst.scrollToItem(lst.currentItem())

        populate()

        def apply_filter(s: str):
            s = (s or "").strip().lower()
            for i in range(lst.count()):
                it = lst.item(i)
                it.setHidden(s not in it.text().lower())

        search.textChanged.connect(apply_filter)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, parent=dlg)
        btn_refresh = QPushButton("Refresh list")
        btns.addButton(btn_refresh, QDialogButtonBox.ActionRole)
        lay.addWidget(btns)

        btn_refresh.clicked.connect(populate)
        lst.itemDoubleClicked.connect(lambda *_: btns.button(QDialogButtonBox.Ok).click())

        def _accept():
            item = lst.currentItem()
            if not item:
                dlg.reject()
                return
            dev_id = int(item.data(Qt.UserRole))
            if kind == "input":
                idx = self.input_device_combo.findData(dev_id)
                if idx >= 0:
                    self.input_device_combo.setCurrentIndex(idx)
            else:
                idx = self.output_device_combo.findData(dev_id)
                if idx >= 0:
                    self.output_device_combo.setCurrentIndex(idx)
            self._update_device_badges()
            dlg.accept()

        btns.accepted.connect(_accept)
        btns.rejected.connect(dlg.reject)

        dlg.exec_()

    def toggle_recording(self, checked=False):
        """Toggle recording; ensure API key; update UI state and timers."""
        if self._is_recording():
            self.record_btn.setEnabled(False)
            self.stop_recording_stage()
            self.record_btn.setText("● Start")
            self.record_btn.setProperty("recording", False)
            self.record_btn.style().unpolish(self.record_btn)
            self.record_btn.style().polish(self.record_btn)
            self.record_btn.setEnabled(True)
            self._restart_volume_monitor()
            self._stop_rec_timer()
        else:
            if not get_api_key():
                self._require_api_key_then(lambda: self.toggle_recording())
                return

            self.record_btn.setEnabled(False)
            self.start_recording_stage()
            self.record_btn.setText("■ Stop")
            self.record_btn.setProperty("recording", True)
            self.record_btn.style().unpolish(self.record_btn)
            self.record_btn.style().polish(self.record_btn)
            self.record_btn.setEnabled(True)
            self._start_rec_timer()

    def eventFilter(self, obj, ev):
        """Overlay meters/captions and clip progress bars under device controls."""
        if obj is self.input_row_frame and hasattr(self, "input_meter"):
            r = self.input_row_frame.rect().adjusted(2, 2, -2, -2)
            self.input_meter.setGeometry(r)

            # Clip meter so it doesn't render beneath the device picker.
            mask = QRegion(self.input_meter.rect())
            target = getattr(self, "input_device_btn", None)
            if not (target and target.isVisible()):
                target = getattr(self, "input_device_combo", None)
            if target and target.isVisible():
                gr = target.geometry()
                gr_in_meter = gr.translated(-r.x(), -r.y()).adjusted(-6, -4, 6, 4)
                mask -= QRegion(gr_in_meter)

            self.input_meter.setMask(mask)

            if hasattr(self, "input_caption"):
                self.input_caption.setGeometry(r)

        elif obj is self.output_row_frame and hasattr(self, "output_meter"):
            r = self.output_row_frame.rect().adjusted(2, 2, -2, -2)
            self.output_meter.setGeometry(r)

            # Clip meter so it doesn't render beneath the device picker.
            mask = QRegion(self.output_meter.rect())
            target = getattr(self, "output_device_btn", None)
            if not (target and target.isVisible()):
                target = getattr(self, "output_device_combo", None)
            if target and target.isVisible():
                gr = target.geometry()
                gr_in_meter = gr.translated(-r.x(), -r.y()).adjusted(-6, -4, 6, 4)
                mask -= QRegion(gr_in_meter)

            self.output_meter.setMask(mask)

            if hasattr(self, "output_caption"):
                self.output_caption.setGeometry(r)

        return QWidget.eventFilter(self, obj, ev)

    def _on_theme_changed(self, text):
        """Switch 'dim'/'night' variant and reapply theme."""
        self._theme_variant = "night" if text.strip().lower() == "night" else "dim"
        self.apply_dark_theme(self._accent, self._theme_variant)
        
    def _refresh_api_key_button_style(self):
        """Danger style when missing key; emerald style when present; resync pill sizes."""
        has = bool(get_api_key())
        self.set_key_btn.setObjectName("recordBtn" if has else "dangerBtn")
        self.set_key_btn.style().unpolish(self.set_key_btn)
        self.set_key_btn.style().polish(self.set_key_btn)
        self._sync_pill_sizes()

    def _select_session(self, which: str):
        """Set mode via big cards, mirror to hidden controls, then refresh sections."""
        self.card_one.setChecked(which == "one")
        self.card_two.setChecked(which == "two")
        self.card_two_aud.setChecked(which == "two_aud")
        self.mode_one.setChecked(which == "one")
        self.mode_two.setChecked(which in ("two", "two_aud"))
        self.chk_audience.setChecked(which in ("one", "two_aud"))
        self._update_sections_from_toggles(which)
        
    def _enter_session_picker(self):
        """Show 3-card picker on Setup tab; clear selection and badge; hide lower panels."""
        if hasattr(self, "tabs") and hasattr(self, "_setup_page"):
            self.tabs.setCurrentWidget(self._setup_page)
        if hasattr(self, "_session_box"):
            self._session_box.setVisible(True)
        if hasattr(self, "session_stack"):
            self.session_stack.setCurrentIndex(0)

        for c in (self.card_one, self.card_two, self.card_two_aud):
            try:
                c.blockSignals(True)
                c.setChecked(False)
                c.blockSignals(False)
            except Exception:
                pass

        try:
            self.mode_one.blockSignals(True)
            self.mode_two.blockSignals(True)
            self.chk_audience.blockSignals(True)
            self.mode_one.setChecked(False)
            self.mode_two.setChecked(False)
            self.chk_audience.setChecked(False)
        finally:
            self.mode_one.blockSignals(False)
            self.mode_two.blockSignals(False)
            self.chk_audience.blockSignals(False)

        if hasattr(self, "session_badge"):
            self.session_badge.setText("Choose session…")
            self.session_badge.setProperty("mode", "")
            self.session_badge.style().unpolish(self.session_badge)
            self.session_badge.style().polish(self.session_badge)

        if hasattr(self, "audience_box"):
            self.audience_box.hide()
        if hasattr(self, "_preset_st_box"):
            self._preset_st_box.hide()
        if hasattr(self, "_preset_ts_box"):
            self._preset_ts_box.hide()

        if hasattr(self, "_setup_page"):
            _s = self._setup_page.layout()
            if _s:
                _s.setStretch(self._i_session, 1)
                _s.setStretch(self._i_audience, 0)
                _s.setStretch(self._i_presets, 0)

        self._last_mode = "none"

    def _enter_session_summary(self, mode: str):
        """Set summary/badge text and QSS mode, collapse picker, expand lower panels."""
        mapping = {
            "one":     "Session: 🎯 One-way (A → B) • audience always present",
            "two":     "Session: 🤝 Two-party (A ↔ B) • no extra audience",
            "two_aud": "Session: 🎤 Two-party + Audience (A ↔ B + 👥)",
        }
        if hasattr(self, "session_summary"):
            self.session_summary.setText(mapping.get(mode, "Session: —"))

        if hasattr(self, "session_badge"):
            badge_map = {
                "one":     "🎯 One-way",
                "two":     "🤝 Two-party",
                "two_aud": "🎤 Two-party + Audience",
            }
            self.session_badge.setText(badge_map.get(mode, "Choose session…"))
            self._sync_pill_sizes()
            self.session_badge.setProperty("mode", mode)
            self.session_badge.style().unpolish(self.session_badge)
            self.session_badge.style().polish(self.session_badge)

        if hasattr(self, "_session_box"):
            self._session_box.setVisible(False)
            
        if hasattr(self, "_setup_page"):
            _s = self._setup_page.layout()
            if _s:
                _s.setStretch(self._i_session, 0)
                _s.setStretch(self._i_audience, 0)
                _s.setStretch(self._i_presets, 1)

    def _populate_voice_combo(self, combo: QComboBox):
        """Fill combo with (label, voice_id); select first item."""
        combo.clear()
        for label, voice_id in VOICE_CHOICES:
            combo.addItem(label, voice_id)  # label shown; id stored in userData
        if combo.count():
            combo.setCurrentIndex(0)

    def record_brief(self):
        """Start/stop brief capture using RecordingThread → TranscriptionThread.

        Guards:
          - Require session mode and API key.
          - Block if main recording is active.
        """
        if self._last_mode == "none":
            self._show_toast("Pick a session first.")
            return
        if not get_api_key():
            self._require_api_key_then(self.record_brief)
            return

        if self.brief_recording_thread and self.brief_recording_thread.isRunning():
            self.append_console("Stopping brief recording…")
            self.brief_recording_thread.stop()
            return

        if self._is_recording():
            self._show_toast("Stop the main recording first.")
            return

        self.append_console("Brief: recording… Speak your setup.")
        self.stop_volume_monitor()  # avoid device contention during capture
        self._set_brief_button_state(True)

        self.brief_recording_thread = RecordingThread(self.fixed_input_device)
        self.brief_recording_thread.progress_signal.connect(self.append_console)
        self.brief_recording_thread.recorded_signal.connect(self._on_brief_recorded)
        self.brief_recording_thread.start()

    def _on_brief_recorded(self, audio_bytes: bytes):
        """Handle brief audio bytes → start transcription."""
        self.append_console(f"Brief: recorded {len(audio_bytes)} bytes.")
        self._set_brief_button_state(False)
        self.start_volume_monitor()
        self._set_bridge("Transcribing...", "busy")

        self.brief_transcription_thread = TranscriptionThread(audio_bytes, tag="Brief")
        self.brief_transcription_thread.progress_signal.connect(self.append_console)
        self.brief_transcription_thread.transcribed_signal.connect(self._on_brief_transcribed)
        self.brief_transcription_thread.start()

    def _on_brief_transcribed(self, text: str):
        """Run brief-filling only if transcript is non-empty."""
        if not text.strip():
            self._set_bridge("Brief: Error", "idle")
            return
        self.append_console(f"Brief: transcript => {text}")
        model = self.gpt_model_input.text().strip() or "gpt-4.1-2025-04-14"
        self._set_bridge("Creating Brief...", "busy")
        self.brief_filling_thread = BriefFillingThread(text, self._last_mode, model)
        self.brief_filling_thread.progress_signal.connect(self.append_console)
        self.brief_filling_thread.filled_signal.connect(self._on_brief_filled)
        self.brief_filling_thread.start()

    def _on_brief_filled(self, data: dict):
        """Apply brief JSON to fields; keep user-entered text."""
        if not data:
            self._show_toast("Couldn't parse brief.")
            return
        self._apply_brief_mapping(self._last_mode, data)
        self._show_toast("Brief fields filled.")
        self._set_bridge("Brief: Ready", "ok")
        self._stop_chip_pulse()

    def _apply_brief_mapping(self, mode: str, data: dict):
        """Fill UI fields from brief; only set empty fields, append notes."""
        def fill_line(w: QLineEdit, val: str):
            if val and not w.text().strip():
                w.setText(val)

        def fill_text(w: QTextEdit, val: str):
            if not val:
                return
            cur = w.toPlainText().strip()
            if not cur:
                w.setPlainText(val)
            else:
                w.setPlainText(cur + ("\n" if not cur.endswith("\n") else "") + val)

        if mode in ("one", "two_aud"):
            fill_line(self.aud_persona, data.get("audience_info", ""))
            fill_line(self.aud_guidance, data.get("audience_expectation", ""))

        st = data.get("st", {})
        if st:
            fill_line(self.st_brief,   st.get("purpose", ""))
            fill_line(self.st_speaker, st.get("speaker", ""))
            fill_line(self.st_tone,    st.get("tone", ""))
            fill_line(self.st_expect,  st.get("expectation", ""))
            fill_text(self.st_notes,   st.get("notes", ""))

        if mode in ("two", "two_aud"):
            ts = data.get("ts", {})
            if ts:
                fill_line(self.ts_brief,   ts.get("purpose", ""))
                fill_line(self.ts_speaker, ts.get("speaker", ""))
                fill_line(self.ts_tone,    ts.get("tone", ""))
                fill_line(self.ts_expect,  ts.get("expectation", ""))
                fill_text(self.ts_notes,   ts.get("notes", ""))

    def _update_sections_from_toggles(self, forced_mode: str = None):
        """Derive mode from toggles (or forced), seed opposite brief if blank, and
        show/hide Audience/TS panels. Also updates summary/badge and labels."""
        if forced_mode in ("one", "two", "two_aud"):
            new_mode = forced_mode
        else:
            if self.mode_one.isChecked():
                new_mode = "one"
            elif self.mode_two.isChecked() and self.chk_audience.isChecked():
                new_mode = "two_aud"
            elif self.mode_two.isChecked():
                new_mode = "two"
            else:
                new_mode = "none"

        prev_mode = getattr(self, "_last_mode", "none")

        def _vals(widgets):
            v = []
            for w in widgets:
                v.append(w.toPlainText().strip() if isinstance(w, QTextEdit) else w.text().strip())
            return v

        def _all_blank(widgets): return all(v == "" for v in _vals(widgets))
        def _any_filled(widgets): return any(v != "" for v in _vals(widgets))

        def _set_if_blank(dest_widgets, src_widgets):
            """Copy src → dest only where dest is blank."""
            svals = _vals(src_widgets)
            for w, v in zip(dest_widgets, svals):
                if not v:
                    continue
                if isinstance(w, QTextEdit):
                    if not w.toPlainText().strip():
                        w.setPlainText(v)
                else:
                    if not w.text().strip():
                        w.setText(v)

        st_widgets = [self.st_brief, self.st_speaker, self.st_tone, self.st_expect, self.st_notes]
        ts_widgets = [self.ts_brief, self.ts_speaker, self.ts_tone, self.ts_expect, self.ts_notes]

        # Seed TS from ST when moving one → (two|two_aud) and TS is blank.
        if prev_mode == "one" and new_mode in ("two", "two_aud") and _all_blank(ts_widgets):
            _set_if_blank(ts_widgets, st_widgets)

        # Seed ST from TS when moving (two|two_aud) → one and ST is blank.
        if prev_mode in ("two", "two_aud") and new_mode == "one" and _any_filled(ts_widgets) and _all_blank(st_widgets):
            _set_if_blank(st_widgets, ts_widgets)

        any_mode = new_mode != "none"
        self._preset_st_box.setVisible(any_mode)
        self._preset_ts_box.setVisible(new_mode in ("two", "two_aud"))
        self.audience_box.setVisible(new_mode in ("one", "two_aud"))
        self._apply_mode_labels(new_mode)

        if new_mode == "none":
            if hasattr(self, "_session_box"):
                self._session_box.setVisible(True)
            if hasattr(self, "session_stack"):
                self.session_stack.setCurrentIndex(0)
        else:
            self._enter_session_summary(new_mode)

        self._last_mode = new_mode

    def start_recording_stage(self):
        """Begin recording stage; stop action timer (recording uses its own)."""
        self._stop_action_timer()
        self._set_status("recording")
        self._set_bridge("REC: Recording…", "busy")
        self._set_input_caption("Recording…")
        self._go_to_review_tab()

        self.append_console("New recording session started.")
        self.stage("REC", "start")
        self.stop_volume_monitor()

        self.recording_thread = RecordingThread(self.fixed_input_device)
        self.recording_thread.progress_signal.connect(self.append_console)
        self.recording_thread.recorded_signal.connect(self.on_recorded)
        self.recording_thread.start()

    def stop_recording_stage(self):
        """End recording stage and clear UI caption."""
        self._set_status("idle")
        if self.recording_thread and self.recording_thread.isRunning():
            self.recording_thread.stop()
            self.recording_thread.wait()
        self.append_console("User stopped recording.")
        self._set_input_caption("")

    def on_recorded(self, audio_bytes: bytes):
        """Kick off ASR after recording completes."""
        self._set_bridge("Transcribing...", "busy")
        self._set_input_caption("Transcribing…")
        self._reset_record_button()
        self.append_console(f"Recorded {len(audio_bytes)} bytes (session total).")
        self.start_transcription_stage(audio_bytes)

    def start_transcription_stage(self, audio_bytes: bytes):
        """Spawn Whisper transcription thread for session audio."""
        self.stage("ASR", "whisper…")
        self.transcription_thread = TranscriptionThread(audio_bytes, tag="Session")
        self.transcription_thread.progress_signal.connect(self.append_console)
        self.transcription_thread.transcribed_signal.connect(self.on_transcribed)
        self.transcription_thread.start()

    def on_transcribed(self, transcript: str):
        """Start translation on ASR result."""
        self._set_bridge("ASR: Done", "ok")
        self._set_input_caption("")
        self.append_console(f"Transcribed => {transcript}")
        self.stage("TR", "translate…")
        self.pending_source_text = transcript
        self.start_translation_stage(transcript)

    def _gather_session_instructions(self) -> str:
        """Collect brief blocks for current mode.

        - one: Audience + ST
        - two: ST + TS
        - two_aud: Audience + ST + TS
        - none: ""
        """
        mode = getattr(self, "_last_mode", "none")
        if mode == "none":
            return ""

        blocks = []

        if mode in ("one", "two_aud"):
            aud_info = self.aud_persona.text().strip()
            aud_expect = self.aud_guidance.text().strip()
            aud_lines = []
            if aud_info:   aud_lines.append(f"Audience Info: {aud_info}")
            if aud_expect: aud_lines.append(f"Audience Expectation: {aud_expect}")
            if aud_lines:
                blocks.append("[AUDIENCE]\n" + "\n".join(aud_lines))

        st_lines = []
        if (v := self.st_brief.text().strip()):    st_lines.append(f"Purpose: {v}")
        if (v := self.st_speaker.text().strip()):  st_lines.append(f"Source Speaker: {v}")
        if (v := self.st_tone.text().strip()):     st_lines.append(f"Tone: {v}")
        if (v := self.st_expect.text().strip()):   st_lines.append(f"Expectation: {v}")
        if (v := self.st_notes.toPlainText().strip()): st_lines.append(f"Notes: {v}")
        if st_lines:
            blocks.append("[ST]\n" + "\n".join(st_lines))

        if mode in ("two", "two_aud"):
            ts_lines = []
            if (v := self.ts_brief.text().strip()):    ts_lines.append(f"Purpose: {v}")
            if (v := self.ts_speaker.text().strip()):  ts_lines.append(f"Target Speaker: {v}")
            if (v := self.ts_tone.text().strip()):     ts_lines.append(f"Tone: {v}")
            if (v := self.ts_expect.text().strip()):   ts_lines.append(f"Expectation: {v}")
            if (v := self.ts_notes.toPlainText().strip()): ts_lines.append(f"Notes: {v}")
            if ts_lines:
                blocks.append("[TS]\n" + "\n".join(ts_lines))

        return "\n\n".join(blocks)

    def start_translation_stage(self, transcript: str):
        """Launch TranslationThread with session brief, context, and term pairs."""
        self._set_bridge("Translating...", "busy")
        source = self.source_lang_input.text().strip()
        target = self.target_lang_input.text().strip()
        
        instr = self._gather_session_instructions()
        model = self.gpt_model_input.text().strip()
        context_text = self.build_translation_context()
        term_lines = self._build_term_lines(False)

        self.append_console(f"Prompt extras → mode={self._last_mode} | ctx={len(context_text)} chars | terms={len([ln for ln in term_lines.splitlines() if ln.strip()])} lines")

        self.translation_thread = TranslationThread(
            transcript, source, target, instr, model, context_text, term_lines
        )
        self.translation_thread.progress_signal.connect(self.append_console)
        self.translation_thread.translated_signal.connect(self.on_translated)
        self.translation_thread.start()

    def on_translated(self, source_text: str, translation: str):
        """Append row to segments table and queue TTS."""
        self.stage("TTSQ", f"{len(self.tts_queue)+1} item(s)")
        self._set_bridge("CCMI: Ready", "ok")
        self.append_console(f"Translated => {translation}")

        row = self.segments_table.rowCount()
        self.segments_table.insertRow(row)

        source_item = QTableWidgetItem(source_text)
        source_item.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
        target_item = QTableWidgetItem(translation)
        target_item.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)

        self.segments_table.setItem(row, 0, source_item)
        self.segments_table.setItem(row, 1, target_item)

        self._update_empty_state()
        self.start_tts_stage(translation)

    def start_tts_stage(self, translation: str):
        """Enqueue translation for TTS; kick off playback if idle."""
        self.tts_queue.append(translation)
        if not self.tts_busy:
            self._dequeue_and_play_tts()

    def on_tts_log(self, msg: str):
        """Update stage chip when playback starts (based on worker logs)."""
        self.append_console(msg)
        if "TTS audio playing" in msg:
            if getattr(self, "_tts_context", "translate") == "translate":
                self._update_stage_chip("play", "Speaking...")
                self._start_chip_pulse(380)
            else:
                self._update_stage_chip("play", "Playing...")
                self._start_chip_pulse(380)

    def on_tts_finished(self):
        """Log end of TTS stage (UI resets elsewhere)."""
        self.append_console("TTS stage finished.")

    def simulate_tts_volume(self):
        pass

    def _apply_mode_labels(self, mode: str):
        """Rename group titles/labels/placeholders per mode and adjust layout."""
        try:
            self.presets_grid.setColumnStretch(0, 1)
            self.presets_grid.setColumnStretch(1, 1)
        except Exception:
            pass

        if mode == "one":
            self.audience_box.setTitle("Audience")
            self.lbl_aud_info.setText("👥 Audience Info:")
            self.lbl_aud_expect.setText("🎯 Audience Expectation:")

            self._preset_st_box.setTitle("Translation Brief")
            self.st_lbl_brief.setText("📝 Purpose:")
            self.st_lbl_speaker.setText("👤 Source Speaker Info:")
            self.st_lbl_tone.setText("🎛️ Tone:")
            self.st_lbl_expect.setText("✅ Source Speaker Expectation:")
            self.st_lbl_notes.setText("📌 Notes / Constraints:")

            self._preset_ts_box.hide()
            self.presets_grid.setColumnStretch(0, 1)
            self.presets_grid.setColumnStretch(1, 0)

            self.aud_persona.setPlaceholderText("Audience (public/staff · ~80)")
            self.aud_guidance.setPlaceholderText("Need (announcements, decisions, next steps)")

            self.st_brief.setPlaceholderText("Purpose (announce/update to audience)")
            self.st_speaker.setPlaceholderText("Source speaker (e.g., Mayor) + style")
            self.st_tone.setPlaceholderText("Tone (clear, formal)")
            self.st_expect.setPlaceholderText("Expectation (faithful & fluent; keep style)")
            self.st_notes.setPlaceholderText("Names/locations; time & numbers")

        elif mode == "two":
            self._preset_st_box.setTitle("Source → Target Translation Brief")
            self.st_lbl_brief.setText("📝 Purpose:")
            self.st_lbl_speaker.setText("👤 Source Speaker Info:")
            self.st_lbl_tone.setText("🎛️ Tone:")
            self.st_lbl_expect.setText("✅ Source Speaker Expectation:")
            self.st_lbl_notes.setText("📌 Notes / Constraints:")

            self._preset_ts_box.setTitle("Target → Source Translation Brief")
            self.ts_lbl_brief.setText("📝 Purpose:")
            self.ts_lbl_speaker.setText("👤 Target Speaker Info:")
            self.ts_lbl_tone.setText("🎛️ Tone:")
            self.ts_lbl_expect.setText("✅ Target Speaker Expectation:")
            self.ts_lbl_notes.setText("📌 Notes / Constraints:")

            self._preset_ts_box.show()
            self.audience_box.hide()
            self.presets_grid.setColumnStretch(0, 1)
            self.presets_grid.setColumnStretch(1, 1)

            self.st_brief.setPlaceholderText("Purpose (Source → Target)")
            self.st_speaker.setPlaceholderText("Source role (e.g., Seller/Manager) + style")
            self.st_tone.setPlaceholderText("Tone to Target (professional, concise)")
            self.st_expect.setPlaceholderText("Expectation (accurate & compact)")
            self.st_notes.setPlaceholderText("Agenda; key numbers; terms to keep")

            self.ts_brief.setPlaceholderText("Purpose (Target → Source)")
            self.ts_speaker.setPlaceholderText("Target role (e.g., Buyer/Client) + style")
            self.ts_tone.setPlaceholderText("Tone to Source (neutral, direct)")
            self.ts_expect.setPlaceholderText("Expectation (faithful; keep nuance)")
            self.ts_notes.setPlaceholderText("Objections; constraints; must-nots")

        elif mode == "two_aud":
            self.audience_box.setTitle("Audience")
            self.lbl_aud_info.setText("👥 Audience Info:")
            self.lbl_aud_expect.setText("🎯 Audience Expectation:")

            self._preset_st_box.setTitle("Source → Target Translation Brief")
            self.st_lbl_brief.setText("📝 Purpose:")
            self.st_lbl_speaker.setText("👤 Source Speaker Info:")
            self.st_lbl_tone.setText("🎛️ Tone:")
            self.st_lbl_expect.setText("✅ Source Speaker Expectation:")
            self.st_lbl_notes.setText("📌 Notes / Constraints:")

            self._preset_ts_box.setTitle("Target → Source Translation Brief")
            self.ts_lbl_brief.setText("📝 Purpose:")
            self.ts_lbl_speaker.setText("👤 Target Speaker Info:")
            self.ts_lbl_tone.setText("🎛️ Tone:")
            self.ts_lbl_expect.setText("✅ Target Speaker Expectation:")
            self.ts_lbl_notes.setText("📌 Notes / Constraints:")

            self._preset_ts_box.show()
            self.audience_box.show()
            self.presets_grid.setColumnStretch(0, 1)
            self.presets_grid.setColumnStretch(1, 1)

            self.aud_persona.setPlaceholderText("Audience (journalists & stakeholders · ~50)")
            self.aud_guidance.setPlaceholderText("Need (on-record quotes, clear takeaways)")

            self.st_brief.setPlaceholderText("Purpose (opening remarks / key points)")
            self.st_speaker.setPlaceholderText("Source speaker (e.g., Minister/Panelist) + style")
            self.st_tone.setPlaceholderText("Tone (on-record, measured)")
            self.st_expect.setPlaceholderText("Expectation (accurate quotes; no ad-lib)")
            self.st_notes.setPlaceholderText("Names/titles; sensitive topics; red lines")

            self.ts_brief.setPlaceholderText("Purpose (questions / replies)")
            self.ts_speaker.setPlaceholderText("Target speaker (e.g., Reporter/Panelist) + style")
            self.ts_tone.setPlaceholderText("Tone (press-friendly, neutral)")
            self.ts_expect.setPlaceholderText("Expectation (faithful quotes; keep nuance)")
            self.ts_notes.setPlaceholderText("Attribution; disclaimers; off-limits")

        else:
            pass

    def _pick_tts_voice(self) -> str:
        """Return voice id based on mode/direction (TS when reversed in two-party)."""
        try:
            if self._last_mode in ("two", "two_aud"):
                combo = self.ts_voice_select if getattr(self, "reverse_mode", False) else self.st_voice_select
            else:
                combo = self.st_voice_select
            return combo.currentData() or "alloy"
        except Exception:
            return "alloy"

    def _test_section_voice(self, which: str):
        """Play a sample sentence for the chosen section's voice (no queue impact)."""
        if not get_api_key():
            self._require_api_key_then(lambda: self._test_section_voice(which))
            return
        if self.tts_busy:
            self._show_toast("TTS is currently playing.")
            return
        sample = "Translate the purpose, not the words."
        combo = self.st_voice_select if which == "st" else self.ts_voice_select
        voice = combo.currentData() or "alloy"
        self._play_tts_once(sample, voice)

    def _play_tts_once(self, text: str, voice: str):
        """One-off TTS playback for 'hear' path; pauses input meter to avoid overlap."""
        self._set_output_caption("Playing…")
        if not get_api_key():
            self._require_api_key_then(lambda: self._play_tts_once(text, voice))
            return
        if self.tts_busy:
            self._show_toast("TTS is currently playing.")
            return

        try:
            self.tts_busy = True
            self.stop_volume_monitor()

            self._tts_context = "hear"
            self._set_bridge("Creating Sound...", "busy")
            self._start_chip_pulse(520)

            self.tts_thread = TTSThread(text, voice, self.fixed_output_device)

            self.tts_thread.progress_signal.connect(self.on_tts_log)
            self.tts_thread.level_signal.connect(self.update_output_meter)
            self.tts_thread.finished.connect(lambda: self._on_tts_done(test=True))
            self.tts_thread.start()
        except Exception as e:
            self.append_console(f"TTS Error: {e}")
            self.tts_busy = False
            self.update_output_meter(0)
            self.start_volume_monitor()

    def _set_brief_button_state(self, recording: bool):
        """Toggle brief buttons and QSS 'recording' state."""
        for btn in (getattr(self, "btn_record_brief_st", None),
                    getattr(self, "btn_record_brief_ts", None)):
            if not btn:
                continue
            btn.setText("■ Stop" if recording else "🎤 Record Brief")
            btn.setProperty("recording", recording)
            btn.setEnabled(True)
            btn.style().unpolish(btn)
            btn.style().polish(btn)

    def append_console(self, text: str):
        """INFO log via root logger (colorlog handles formatting)."""
        logging.getLogger().info(text)

    def _go_to_review_tab(self):
        """Activate Review tab (uses stored widget; falls back to title scan)."""
        try:
            if hasattr(self, "tabs"):
                if getattr(self, "_review_page", None) is not None:
                    self.tabs.setCurrentWidget(self._review_page)
                    return
                for i in range(self.tabs.count()):
                    if self.tabs.tabText(i).lower().startswith("review"):
                        self.tabs.setCurrentIndex(i)
                        return
        except Exception:
            pass

    def _update_stage_chip(self, state: str, label: str):
        """Set chip state/label; show logo only in idle; drive action timer when not recording."""
        if hasattr(self, "stage_chip") and self.stage_chip:
            try:
                if not hasattr(self, "_idle_logo"):
                    self._idle_logo = None
                    for name in ("ccmi_logo.png", "logo.png"):
                        p = resource_path(name)
                        if os.path.exists(p):
                            self._idle_logo = QPixmap(p)
                            break
            except Exception:
                self._idle_logo = None

            use_logo = (
                state == "idle"
                and getattr(self, "_idle_logo", None) is not None
                and not self._idle_logo.isNull()
            )
            if use_logo:
                h = max(16, self.stage_chip.height() - 6)
                self.stage_chip.setPixmap(self._idle_logo.scaledToHeight(h, Qt.SmoothTransformation))
                self.stage_chip.setText("")
                self.stage_chip.setToolTip("Idle")
            else:
                self.stage_chip.setPixmap(QPixmap())
                self.stage_chip.setText(label or "")
                self.stage_chip.setToolTip(label or "")

            self.stage_chip.setProperty("state", state)
            self.stage_chip.style().unpolish(self.stage_chip)
            self.stage_chip.style().polish(self.stage_chip)

        if not self._is_recording():
            if state in ("busy", "play"):
                self._start_action_timer()
            elif state in ("idle", "ok"):
                self._stop_action_timer()

    def stage(self, key: str, detail: str = ""):
        """Stage logger + chip update.

        key ∈ {'REC','ASR','TR','TTSQ','TTS','OK','ERR'}; maps to chip state/pulse.
        """
        em = {
            "REC":"🎙️", "ASR":"📝", "TR":"🌐",
            "TTSQ":"⏳", "TTS":"🔊", "OK":"✅", "ERR":"⛔"
        }.get(key, "•")
        tag = f"[{key}]"
        msg = f"{em} {tag} {detail}".strip()

        self.append_console(msg)

        state = {
            "REC":"busy", "ASR":"busy", "TR":"busy",
            "TTSQ":"busy", "TTS":"play", "OK":"ok", "ERR":"idle"
        }.get(key, "idle")
        short = {
            "REC": f"{em} Rec…",
            "ASR": f"{em} ASR…",
            "TR":  f"{em} TR…",
            "TTSQ":f"{em} Queue",
            "TTS": f"{em} Playing",
            "OK":  f"{em} Done",
            "ERR": f"{em} Error",
        }.get(key, f"{em} {key}")
        self._update_stage_chip(state, short)

        if state in ("busy", "play"):
            self._start_chip_pulse(520 if state == "busy" else 380)
        else:
            self._stop_chip_pulse()

    def _start_chip_pulse(self, interval_ms: int = 500):
        """Blink chip by toggling a 'pulse' property at a fixed interval."""
        if not hasattr(self, "_chip_pulse_timer"):
            self._chip_pulse_timer = QTimer(self)
            self._chip_pulse_timer.timeout.connect(self._flip_chip_pulse)
        self._chip_pulse_on = False
        self._chip_pulse_timer.start(interval_ms)

    def _stop_chip_pulse(self):
        """Stop pulsing and ensure property is off."""
        if hasattr(self, "_chip_pulse_timer"):
            self._chip_pulse_timer.stop()
        self._set_chip_pulse(False)

    def _flip_chip_pulse(self):
        """Toggle pulse flag and apply to chip."""
        self._chip_pulse_on = not getattr(self, "_chip_pulse_on", False)
        self._set_chip_pulse(self._chip_pulse_on)

    def _set_chip_pulse(self, on: bool):
        """Set 'pulse' property to drive QSS animation."""
        if hasattr(self, "stage_chip") and self.stage_chip:
            self.stage_chip.setProperty("pulse", "on" if on else "off")
            self.stage_chip.style().unpolish(self.stage_chip)
            self.stage_chip.style().polish(self.stage_chip)

    def _bridge_langs_tag(self) -> str:
        """Return 'Source→Target' tag based on current inputs."""
        return f"{self.source_lang_input.text().strip()}→{self.target_lang_input.text().strip()}"

    def _set_bridge(self, text: str, state: str = "idle"):
        """Update chip with a transient bridge label/state; logo only in idle; drives action timer."""
        if hasattr(self, "stage_chip"):
            try:
                if not hasattr(self, "_idle_logo"):
                    self._idle_logo = None
                    for name in ("ccmi_logo.png", "logo.png"):
                        p = resource_path(name)
                        if os.path.exists(p):
                            self._idle_logo = QPixmap(p)
                            break
            except Exception:
                self._idle_logo = None

            use_logo = (
                state == "idle"
                and getattr(self, "_idle_logo", None) is not None
                and not self._idle_logo.isNull()
            )
            if use_logo:
                h = max(16, self.stage_chip.height() - 6)
                self.stage_chip.setPixmap(self._idle_logo.scaledToHeight(h, Qt.SmoothTransformation))
                self.stage_chip.setText("")
                self.stage_chip.setToolTip(text or "Idle")
            else:
                self.stage_chip.setPixmap(QPixmap())
                self.stage_chip.setText(text or "")
                self.stage_chip.setToolTip(text or "")

            self.stage_chip.setProperty("state", state)
            self.stage_chip.style().unpolish(self.stage_chip)
            self.stage_chip.style().polish(self.stage_chip)

        if not self._is_recording():
            if state in ("busy", "play"):
                self._start_action_timer()
            elif state in ("idle", "ok"):
                self._stop_action_timer()

    def _set_input_caption(self, text: str):
        """Set input row stage ('idle' or 'busy') to drive QSS tint; hides overlay."""
        if hasattr(self, "input_caption"):
            self.input_caption.hide()

        state = "idle" if not text else "busy"
        try:
            self.input_row_frame.setProperty("stage", state)
            self.input_row_frame.style().unpolish(self.input_row_frame)
            self.input_row_frame.style().polish(self.input_row_frame)

            if hasattr(self, "_dev_card"):
                self._dev_card.setProperty("stage", state)
                self._dev_card.style().unpolish(self._dev_card)
                self._dev_card.style().polish(self._dev_card)
        except Exception:
            pass

    def _set_output_caption(self, text: str):
        """Set output row stage: 'play' when text contains 'play', else 'busy'/idle; hides overlay."""
        if hasattr(self, "output_caption"):
            self.output_caption.hide()

        lower = (text or "").lower()
        state = "idle" if not text else ("play" if "play" in lower else "busy")
        try:
            self.output_row_frame.setProperty("stage", state)
            self.output_row_frame.style().unpolish(self.output_row_frame)
            self.output_row_frame.style().polish(self.output_row_frame)

            if hasattr(self, "_dev_card"):
                self._dev_card.setProperty("stage", state if state in ("busy", "play") else "idle")
                self._dev_card.style().unpolish(self._dev_card)
                self._dev_card.style().polish(self._dev_card)
        except Exception:
            pass

    def _show_toast(self, msg: str, ms: int = 2200):
        """Transient toast in top-right; auto-hides via single-shot timer."""
        self.toast.setText(msg)
        self.toast.adjustSize()
        self.toast.move(self.width() - self.toast.width() - 24, 18)
        self.toast.show()
        self._toast_timer.start(ms)

    def _hide_toast(self):
        """Hide toast immediately."""
        self.toast.hide()

    def build_translation_context(self) -> str:
        """Concatenate prior translations and cap to last 5000 chars for prompt size."""
        tcol = self._get_translation_col_index()
        parts = []
        for r in range(self.segments_table.rowCount()):
            item = self.segments_table.item(r, tcol)
            if item:
                t = item.text().strip()
                if t:
                    parts.append(t)
        ctx = "\n\n".join(parts)
        return ctx[-5000:]

    def _get_translation_col_index(self) -> int:
        """Find 'Translation' column index; default to 1."""
        for c in range(self.segments_table.columnCount()):
            header = self.segments_table.horizontalHeaderItem(c)
            if header and header.text().strip().lower() == "translation":
                return c
        return 1

    def _dequeue_and_play_tts(self):
        """Pop next TTS item; skip if muted; set chip/output states; start thread."""
        if not self.tts_queue:
            return
        self.tts_busy = True
        text = self.tts_queue.pop(0)
        if self.tts_muted:
            self.append_console("TTS muted; skipping playback.")
            self._on_tts_done()
            return
        voice_name = self._pick_tts_voice()

        self._tts_context = "translate"
        self._set_bridge("Interpreting...", "busy")
        self._set_output_caption("Playing…")

        self.tts_thread = TTSThread(text, voice_name, self.fixed_output_device)

        self.tts_thread.progress_signal.connect(self.on_tts_log)
        self.tts_thread.level_signal.connect(self.update_output_meter)
        self.tts_thread.finished.connect(lambda: self._on_tts_done(test=False))
        self.tts_thread.start()

    def _on_tts_done(self, test: bool = False):
        """Cleanup after TTS; restore chip; continue queue or restart input monitor."""
        self._set_output_caption("")
        if test:
            self._stop_chip_pulse()
            if self._chip_restore:
                prev_text, prev_state = self._chip_restore
                self._update_stage_chip(prev_state or "idle", prev_text or "Idle")
                self._chip_restore = None
            else:
                self._update_stage_chip("idle", "Idle")
        else:
            self.stage("OK", "segment done")
            self._update_stage_chip("idle", "Idle")

        self.on_tts_finished()
        self.tts_busy = False
        self.update_output_meter(0)
        if self.tts_queue:
            self._dequeue_and_play_tts()
        else:
            self.start_volume_monitor()

    def _reset_record_button(self):
        """Return Start button to idle style/state after recording."""
        self.record_btn.setText("● Start")
        self.record_btn.setProperty("recording", False)
        self.record_btn.style().unpolish(self.record_btn)
        self.record_btn.style().polish(self.record_btn)

    def update_output_meter(self, vol: int):
        """Set output progress bar to current level (0–100)."""
        if hasattr(self, "output_meter"):
            self.output_meter.setValue(vol)

    def _copy_last_translation(self):
        """Copy newest translation cell to clipboard; toasts on empty state."""
        row = self.segments_table.rowCount() - 1
        if row < 0:
            self._show_toast("No segments yet.")
            return
        col = self._get_translation_col_index()
        item = self.segments_table.item(row, col)
        if not item:
            self._show_toast("Nothing to copy.")
            return
        QApplication.clipboard().setText(item.text())
        self._show_toast("Last translation copied.")

    def _export_xlsx(self):
        """Export table to XLSX with wrapped, top-aligned cells and autosized columns."""
        path, _ = QFileDialog.getSaveFileName(self, "Export Segments", "segments.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Font

            wb = Workbook()
            ws = wb.active
            ws.title = "CCMI Segments"

            headers = [self.segments_table.horizontalHeaderItem(c).text()
                       for c in range(self.segments_table.columnCount())]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical="center")

            for r in range(self.segments_table.rowCount()):
                row = []
                for c in range(self.segments_table.columnCount()):
                    item = self.segments_table.item(r, c)
                    row.append("" if item is None else item.text())
                ws.append(row)

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            for col_idx in range(1, ws.max_column + 1):
                max_len = 0
                for row_idx in range(1, ws.max_row + 1):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is None:
                        continue
                    s = str(v)
                    line_len = max((len(line) for line in s.splitlines()), default=0)
                    max_len = max(max_len, line_len)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

            ws.freeze_panes = "A2"

            wb.save(path)
            self._show_toast("Exported XLSX.")
        except ImportError:
            self._show_toast("openpyxl not installed; cannot export XLSX.")
        except Exception as e:
            self._show_toast(f"Export error: {e}")

    def _clear_segments(self):
        """Clear Review table and refresh empty state."""
        self.segments_table.setRowCount(0)
        self._update_empty_state()
        self._show_toast("Segments cleared.")

    def _update_empty_state(self):
        """Placeholder for future empty-state UX."""
        pass

    def _toggle_help(self):
        """Reserved for help overlay toggle."""
        pass


if __name__ == "__main__":
    import traceback

    if os.name == "nt":
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                u"CCMI.Interpreter"
            )
        except Exception:
            pass

    app = QApplication(sys.argv)

    try:
        ico = resource_path("app.ico")
        if os.path.exists(ico):
            app.setWindowIcon(QIcon(ico))
    except Exception:
        pass

    try:
        w = MainWindow()
        try:
            ico = resource_path("app.ico")
            if os.path.exists(ico):
                w.setWindowIcon(QIcon(ico))
        except Exception:
            pass

        w.show()
        sys.exit(app.exec_())
    except Exception as e:
        print("Fatal error:", e)
        traceback.print_exc()
        sys.exit(1)

